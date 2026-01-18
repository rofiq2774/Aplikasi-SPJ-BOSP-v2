import os
import io
import httpx
import sys
import subprocess
import logging
import fitz  
import re
import shutil
import json
import uuid
import openpyxl
from urllib.parse import urlparse
from PyPDF2 import PdfMerger
from urllib.parse import unquote
from PIL import Image
from alembic.config import Config
from alembic import command
from urllib.parse import quote
from fpdf import FPDF
from io import BytesIO
from sqlalchemy import insert, delete, update, select, and_, or_, func
from sqlalchemy.ext.asyncio import AsyncSession
from dateutil.parser import parse as parse_date
from typing import Dict, List, Optional
from collections import defaultdict
from pathlib import Path
from datetime import datetime, timezone
from reportlab.lib.utils import simpleSplit
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm, mm
from num2words import num2words
from pydantic import BaseModel
from fastapi import (
    FastAPI, APIRouter, UploadFile, File,
    HTTPException, Depends, Form, Request, Query
)
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv
from contextlib import asynccontextmanager
from database import engine, get_db, Base, BASE_DIR
from models import (
    Pengaturan,
    MasterKegiatan,
    Transaksi,
    Kwitansi,
    MasterRekeningBelanja
)
_migrated = False

# ===============================
# ENV & MODE
# ===============================
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

IS_DEV = os.getenv("PYTHON_ENV") == "development"

# ===============================
# LOGGING (SATU KALI SAJA)
# ===============================
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
log_level = getattr(logging, LOG_LEVEL, logging.INFO)

logging.basicConfig(
    level=log_level,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)

logger = logging.getLogger(__name__)

# Matikan log SQLAlchemy di production
if not IS_DEV:
    logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)
    logging.getLogger("sqlalchemy.pool").setLevel(logging.WARNING)
    logging.getLogger("sqlalchemy.dialects").setLevel(logging.WARNING)

log_file = BASE_DIR / "app.log"

file_handler = logging.FileHandler(log_file, encoding="utf-8")
file_handler.setLevel(log_level)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
file_handler.setFormatter(formatter)

# 1. Tambahkan ke Root Logger (untuk aplikasi ini & library umum)
root_logger = logging.getLogger()
root_logger.addHandler(file_handler)

# 2. Tambahkan khsus ke Uvicorn (agar log request HTTP & Startup masuk file)
# Gunakan try-except karena logging uvicorn mungkin belum terinit jika dijalankan manual
try:
    logging.getLogger("uvicorn").addHandler(file_handler)
    logging.getLogger("uvicorn.error").addHandler(file_handler)
    logging.getLogger("uvicorn.access").addHandler(file_handler)
except Exception:
    pass

logger.info(f"Log file aktif di: {log_file}")

# ===============================
# FASTAPI APP (with lifespan)
# ===============================
@asynccontextmanager
async def lifespan(app):
    logger.info("Inisialisasi database...")
    try:
        run_migrations()
    except Exception as e:
        logger.exception("run_migrations gagal: %s", e)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    logger.info("Database siap")
    yield

app = FastAPI(title="Aplikasi SPJ BOSP", lifespan=lifespan)

# ===============================
# HEALTH CHECK (UNTUK ELECTRON)
# ===============================
@app.get("/health")
async def health_check():
    return {
        "status": "siap",
        "waktu": datetime.now(timezone.utc)
    }

# ===============================
# CORS
# ===============================
if IS_DEV:
    cors_origins = [
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost",
        "http://127.0.0.1",
    ]
else:
    cors_origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logger.info(f"CORS allow_origins={cors_origins}")
# ===============================
# DIRECTORIES (PAKAI BASE_DIR DARI database.py)
# ===============================
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_LOGO_DIR = UPLOAD_DIR / "logos"
UPLOAD_KWITANSI_DIR = UPLOAD_DIR / "kwitansi"
UPLOAD_LOGO_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_KWITANSI_DIR.mkdir(parents=True, exist_ok=True)


# ===============================
# STATIC FILES
# ===============================
app.mount(
    "/uploads",
    StaticFiles(directory=str(UPLOAD_DIR)),
    name="uploads"
)
# ===============================
# ROUTER API
# ===============================
api_router = APIRouter(prefix="/api")

# (endpoint-endpoint Anda lanjutkan di bawah sini)

app.include_router(api_router)

# Definisi Class
class AIImageRequest(BaseModel):
    prompt_items: str
    jenis_foto: str
# Pydantic Models
class TransaksiItem(BaseModel):
    tanggal: str
    kode_kegiatan: Optional[str] = ""
    kode_rekening: Optional[str] = ""
    no_bukti: str
    uraian: str
    nama_kegiatan: Optional[str] = ""
    penerimaan: float = 0
    pengeluaran: float = 0
    saldo: float = 0
    volume: float = 0
    satuan: Optional[str] = ""
    harga_satuan: float = 0.0


class BKUData(BaseModel):
    no_bukti: Optional[str] = ""
    transaksi: List[TransaksiItem] = []


class KwitansiRequest(BaseModel):
    transaksi: TransaksiItem
    nama_sekolah: Optional[str] = ""
    alamat_sekolah: Optional[str] = ""
    bulan: Optional[str] = ""
    tahun: Optional[str] = ""
    format: str = "pdf"


class PengaturanRequest(BaseModel):
    nama_sekolah: str = ""
    nama_kepala_sekolah: str = ""
    nip_kepala_sekolah: str = ""
    nama_bendahara: str = ""
    nip_bendahara: str = ""
    nama_pengurus_barang: str = ""
    nip_pengurus_barang: str = ""    
    alamat_sekolah: str = ""
    tempat_surat: str = ""


class MasterKegiatanRequest(BaseModel):
    kode_kegiatan: str
    nama_kegiatan: str


class MasterKegiatanBulkRequest(BaseModel):
    data: List[MasterKegiatanRequest]


class MasterRekeningBelanjaRequest(BaseModel):
    kode_rekening_belanja: str
    nama_rekening_belanja: str
    rekap_rekening_belanja: Optional[str] = ""
    nilai_kapitalisasi_belanja: Optional[float] = 0.0


class MasterRekeningBelanjaBulkRequest(BaseModel):
    data: List[MasterRekeningBelanjaRequest]

def run_migrations():
    global _migrated
    if _migrated:
        logger.info("Migration sudah pernah dijalankan, skip.")
        return
    _migrated = True

    logger.info("Menjalankan auto migration...")

    try:
        # ===============================
        # 1. DETEKSI PATH (UNTUK EXE)
        # ===============================
        if getattr(sys, 'frozen', False):
            # Jika berjalan sebagai .exe (PyInstaller)
            base_path = Path(sys._MEIPASS)
        else:
            # Jika berjalan sebagai script .py biasa
            base_path = Path(__file__).parent

        alembic_ini = base_path / "alembic.ini"
        alembic_dir = base_path / "alembic"

        # Pastikan database diarahkan ke lokasi permanen (AppData), bukan temp folder
        db_path = BASE_DIR / "aplikasi_spj.db"
        db_url = f"sqlite:///{db_path}"

        if not alembic_ini.exists() or not alembic_dir.exists():
            logger.warning("File atau folder Alembic tidak ditemukan. Melewati migrasi.")
            return

        # ===============================
        # 2. KONFIGURASI ALEMBIC
        # ===============================
        cfg = Config(str(alembic_ini))
        cfg.set_main_option("script_location", str(alembic_dir))
        cfg.set_main_option("sqlalchemy.url", db_url)

        from alembic.runtime.migration import MigrationContext
        from sqlalchemy import create_engine
        from alembic.script import ScriptDirectory

        engine_sync = create_engine(db_url)
        script = ScriptDirectory.from_config(cfg)

        with engine_sync.connect() as conn:
            context = MigrationContext.configure(conn)
            current_rev = context.get_current_revision()

        logger.info(f"Current DB revision: {current_rev}")

        heads = script.get_heads()
        head = heads[0] if heads else None

        if not head:
            logger.warning("Tidak ada revision head ditemukan di folder versions.")
            return

        # ===============================
        # 3. STRATEGI UPGRADE (HARDENING)
        # ===============================
        if current_rev is None:
            # Kondisi: Database baru atau tabel alembic_version kosong
            logger.warning("DB belum memiliki revision. Mencoba menjalankan upgrade ke head...")
            try:
                # Upgrade akan mencoba membuat tabel atau menambah kolom yang kurang
                command.upgrade(cfg, "head")
                logger.info("Upgrade database awal berhasil.")
            except Exception as e:
                # Jika upgrade gagal (misal tabel sudah ada tapi manual), kita paksa stamp
                logger.error(f"Upgrade gagal: {e}. Melakukan stamp sebagai fallback.")
                command.stamp(cfg, "head")
        else:
            # Kondisi: Database sudah punya nomor versi
            try:
                # Validasi apakah revisi di DB dikenali oleh script yang dibawa EXE
                script.get_revision(current_rev)
                logger.info(f"Revisi {current_rev} valid. Memastikan database up-to-date...")
                command.upgrade(cfg, "head")
            except Exception as e:
                logger.warning(f"Revision mismatch atau error: {e}. Memaksa sinkronisasi...")
                # Jika revisi DB tidak dikenal (mungkin karena ganti file migrasi),
                # kita tandai yang sekarang sebagai yang terbaru lalu upgrade.
                command.stamp(cfg, "head")
                command.upgrade(cfg, "head")

        logger.info("Auto migration selesai dengan sukses.")

    except Exception:
        # Logging exception tanpa menghentikan startup aplikasi utama
        logger.exception("Terjadi kesalahan fatal pada proses migrasi")


def resolve_foto_path(foto_rel_path: str) -> Path | None:
    if not foto_rel_path:
        return None

    try:
        p = Path(foto_rel_path)

        # absolute langsung
        if p.is_absolute() and p.exists():
            return p

        # dari BASE_DIR
        p2 = Path(BASE_DIR) / foto_rel_path
        if p2.exists():
            return p2

        # strip slash
        p3 = Path(BASE_DIR) / foto_rel_path.lstrip("/\\")
        if p3.exists():
            return p3

        logger.error(f"[FOTO] Tidak ditemukan: {foto_rel_path}")
        return None

    except Exception:
        logger.exception(f"[FOTO] Error resolve: {foto_rel_path}")
        return None

def resolve_logo(path_input: str) -> str:
    """
    Menerima:
    - URL: http://localhost:8000/uploads/logos/a.png
    - Relatif: uploads/logos/a.png
    - Nama file: a.png

    Mengembalikan:
    - Absolute filesystem path
    """

    if not path_input:
        return ""

    try:
        # Kalau berupa URL → ambil path-nya saja
        if path_input.startswith("http://") or path_input.startswith("https://"):
            parsed = urlparse(path_input)
            path_input = parsed.path  # /uploads/logos/a.png

        p = Path(path_input)

        # 1. Kalau sudah absolute dan ada
        if p.is_absolute() and p.exists():
            return str(p)

        # 2. Dari BASE_DIR
        p2 = Path(BASE_DIR) / path_input.lstrip("/")
        if p2.exists():
            return str(p2)

        # 3. Dari uploads/logos
        p3 = UPLOAD_LOGO_DIR / p.name
        if p3.exists():
            return str(p3)

        logger.error(f"Logo tidak ditemukan: {path_input}")
        return ""

    except Exception as e:
        logger.exception(f"Error resolve_logo: {path_input} -> {e}")
        return ""


def extract_bulan_from_tanggal(tanggal: str) -> str:
    """Extract month from date string (supports DD/MM/YYYY or DD-MM-YYYY)"""
    if not tanggal:
        return "1"
    match = re.search(r'\d{1,2}[-/](\d{1,2})[-/]\d{2,4}', tanggal)
    if match:
        return match.group(1)
    
    return "1"


def extract_tahun_from_tanggal(tanggal: str) -> str:
    """Extract year from date string"""
    if not tanggal:
        return str(datetime.now().year)
    
    match = re.search(r'(\d{4})', tanggal)
    if match:
        return match.group(1)
    
    return str(datetime.now().year)

async def get_nama_kegiatan_from_kode(
    kode_kegiatan: str,
    db: AsyncSession
) -> str:
    if not kode_kegiatan:
        return ""

    stmt = select(MasterKegiatan.nama_kegiatan)\
        .where(MasterKegiatan.kode_kegiatan == kode_kegiatan)

    result = await db.execute(stmt)
    return result.scalar() or ""

# --- SETUP LOGGING ---
logger = logging.getLogger(__name__)

# --- UPDATE KONFIGURASI INI ---

# Naikkan toleransi Y agar teks yang sedikit tidak rata tetap dianggap satu baris
Y_TOLERANCE = 10  # Sebelumnya 5, naikkan jadi 10-12

COLUMNS = {
    "tanggal": (30, 85),
    "kode_kegiatan": (85, 140),
    "kode_rekening": (130, 198), 
    "no_bukti": (195, 270), 
    "uraian": (240, 610),      # Tetap mulai dari 240 untuk menangkap "Saldo Bank"
    "penerimaan": (610, 720),
    "pengeluaran": (720, 790),
    "saldo": (790, 950),
}

# --- HELPER FUNCTIONS ---

def parse_amount(val: str) -> float:
    if not val: return 0.0
    # Hilangkan Rp, spasi, dan titik ribuan
    clean_val = str(val).replace("Rp", "").replace(" ", "").replace(".", "")
    # Ganti koma desimal menjadi titik
    clean_val = clean_val.replace(",", ".")
    
    # Ambil hanya karakter yang valid untuk angka (termasuk tanda negatif)
    clean_val = re.sub(r"[^0-9.\-]", "", clean_val)
    
    try:
        return float(clean_val) if clean_val else 0.0
    except:
        return 0.0

def detect_column(x: float):
    """Menentukan kolom berdasarkan koordinat X"""
    for name, (xmin, xmax) in COLUMNS.items():
        if xmin <= x < xmax:
            return name
    return None

def extract_lines_with_position(page):
    """Mengambil teks dan koordinat dari PDF"""
    blocks = page.get_text("dict")["blocks"]
    lines = []
    for block in blocks:
        if block["type"] != 0: continue
        for line in block["lines"]:
            for span in line["spans"]:
                text = span["text"].strip()
                if not text: continue
                x0, y0, x1, y1 = span["bbox"]
                # Debug output moved to logger; keep available in dev or when LOG_TO_TERMINAL=true
                if IS_DEV:
                    logger.debug(f"DEBUG: Teks '{text:30}' | X: {x0:7.2f} | Y: {y0:7.2f}") #jangan dihapus!
                lines.append({
                    "text": text,
                    "x": x0,
                    "y": y0,
                    "height": y1 - y0
                })
    return lines

def group_lines_into_rows(lines, y_tolerance=Y_TOLERANCE):
    """Menggabungkan teks yang berada di baris (Y) yang sama"""
    lines.sort(key=lambda item: item["y"])
    rows = []
    current_row = {}
    last_y = -100

    for item in lines:
        col_name = detect_column(item["x"])
        if not col_name: continue

        if abs(item["y"] - last_y) > y_tolerance:
            if current_row: rows.append(current_row)
            current_row = defaultdict(str)
            last_y = item["y"]
        
        if current_row[col_name]:
            current_row[col_name] += " " + item["text"]
        else:
            current_row[col_name] = item["text"]
            
    if current_row: rows.append(current_row)
    return rows

def print_debug_table(transaksi_list):
    """Tampilan tabel terminal yang disesuaikan dengan kolom BKU di gambar.
    Fungsi ini hanya menulis output ketika `IS_DEV` True untuk membantu debugging
    ekstraksi PDF (menggunakan `logger.debug`)."""
    header = (f"{'TANGGAL':<12} | {'KODE KEG':<12} | {'KODE REK':<15} | {'NO BUKTI':<12} | "
              f"{'URAIAN (Potongan)':<40} | {'MASUK':>12} | {'KELUAR':>12} | {'SALDO':>12}")

    if not transaksi_list:
        if IS_DEV:
            logger.debug("DEBUG: Tidak ada transaksi untuk ditampilkan")
        return

    if IS_DEV:
        logger.debug(header)
        logger.debug('-' * len(header))
        for t in transaksi_list:
            # Ambil atribut dengan safe-get agar fungsi tidak fail jika struktur berubah
            tanggal = getattr(t, 'tanggal', '') or ''
            kode_keg = getattr(t, 'kode_kegiatan', '') or '-'
            kode_rek = getattr(t, 'kode_rekening', '') or '-'
            no_bukti = getattr(t, 'no_bukti', '') or '-'
            uraian = getattr(t, 'uraian', '') or ''
            u_short = (uraian[:37] + '...') if len(uraian) > 40 else uraian
            masuk = getattr(t, 'penerimaan', 0) or 0
            keluar = getattr(t, 'pengeluaran', 0) or 0
            saldo = getattr(t, 'saldo', 0) or 0

            line = (f"{tanggal:<12} | {kode_keg:<12} | {kode_rek:<15} | {no_bukti:<12} | "
                    f"{u_short:<40} | {masuk:12.2f} | {keluar:12.2f} | {saldo:12.2f}")

            logger.debug(line)
    # Jika bukan mode development, jangan output table
    return

# --- MAIN EXTRACTION FUNCTION ---
async def extract_bku_data(pdf_content: bytes, db):
    doc = fitz.open(stream=pdf_content, filetype="pdf")
    transaksi_list = []

    BLACKLIST = [
        "BUKU KAS UMUM", "BUKUKASUMUM", "NPSN", "NAMA SEKOLAH", "NAMASEKOLAH", 
        "MENYETUJUI", "KEPALA SEKOLAH", "KEPALASEKOLAH", "BENDAHARA", 
        "DITUTUP DENGAN", "DITUTUPDENGAN", "HALAMAN", "NIP.", 
        "TOTAL PENERIMAAN", "TOTAL PENGELUARAN"
    ]
    
    # Keyword khusus untuk footer baris jumlah/total di bawah tabel
    FOOTER_KEYWORDS = ["JUMLAH", "TOTAL", "SISA"]

    bpu_pattern = re.compile(r'(\d*BPU\s*\d+)', re.IGNORECASE)
    # Update regex tanggal agar support spasi (jika OCR kurang pas) e.g 03 09 2025
    date_pattern = re.compile(r'\d{2}[-/\s]\d{2}[-/\s]\d{4}')

    try:
        last_valid_date = ""

        for page in doc:
            words = page.get_text("words")
            lines = {}
            
            # --- LOGIKA GROUPING BARIS ---
            for w in words:
                y_coord = round(w[1])
                found_line = False
                for existing_y in list(lines.keys()):
                    if abs(y_coord - existing_y) <= 8: # Tolerance y
                        lines[existing_y].append(w)
                        found_line = True
                        break
                if not found_line:
                    lines[y_coord] = [w]

            for y in sorted(lines.keys()):
                line_words = sorted(lines[y], key=lambda x: x[0])
                full_line_text = " ".join([w[4] for w in line_words]).strip()
                
                if not full_line_text:
                    continue

                clean_text_check = full_line_text.replace(" ", "").upper()

                is_date_start = date_pattern.search(full_line_text[:12])
                
                if not is_date_start:
                     if any(k in clean_text_check for k in BLACKLIST):
                        continue
                     # Cek footer khusus (Jumlah/Total)
                     if any(full_line_text.upper().startswith(k) for k in FOOTER_KEYWORDS):
                         continue

                # Filter baris nomor kolom (hanya angka 1 s.d 8)
                if re.fullmatch(r'[\d\s\.]+', full_line_text):
                    continue

                tgl, bukti = "", ""
                urai, kgt_raw, rek_raw = [], [], []
                masuk, keluar, sld = 0, 0, 0

                # --- PARSING KOLOM (KOORDINAT DISESUAIKAN) ---
                for w in line_words:
                    x0, text = w[0], w[4].strip()
                    if not text: continue

                    # Cek tanggal duluan sebelum cek koordinat
                    if not tgl and date_pattern.search(text):
                        tgl = text
                        continue
                    if x0 < 80: pass 
                    elif 80 <= x0 < 132: kgt_raw.append(text)    # Kode Kegiatan
                    elif 132 <= x0 < 195: rek_raw.append(text)   # Kode Rekening
                    elif 195 <= x0 < 240:                        # Bukti / Awal Uraian
                        if bpu_pattern.search(text): bukti = text
                        else: urai.append(text)
                    elif 240 <= x0 < 600: urai.append(text)      # Uraian
                    elif 600 <= x0 < 675:                        # Penerimaan
                        val = parse_amount(text)
                        if val: masuk = val
                    elif 675 <= x0 < 755:                        # Pengeluaran
                        val = parse_amount(text)
                        if val: keluar = val
                    elif x0 >= 755:                              # Saldo
                        val = parse_amount(text)
                        if val: sld = val

                uraian_final = " ".join(urai).strip()
                
                # --- LOGIKA VALIDASI BARIS ---
                is_financial_row = (masuk > 0 or keluar > 0)
                # Baris dianggap konten jika punya Tanggal ATAU Bukti ATAU Kode Rekening
                has_content = (tgl != "" or bukti != "" or len(rek_raw) > 0)
                
                # Header kolom (seperti "URAIAN", "NO BUKTI") biasanya tidak punya uang & tidak punya kode rekening valid
                if not is_financial_row and not has_content:
                    # Cek tambahan: jika uraian sangat pendek/kosong, skip
                    if len(uraian_final) < 3: 
                        continue
                    # Jika hanya teks header kolom, skip
                    if "URAIAN" in uraian_final.upper() or "BUKTI" in uraian_final.upper():
                        continue

                # Pewarisan Tanggal
                if tgl:
                    last_valid_date = tgl
                elif last_valid_date and (is_financial_row or bukti or len(uraian_final) > 3):
                    # Wariskan tanggal jika ini terlihat seperti lanjutan baris transaksi
                    tgl = last_valid_date
                
                if not tgl:
                    continue

                # --- CLEANING KODE ---
                full_rek_raw = "".join(rek_raw).replace(" ", "")
                match_rek = re.search(r'(\d+(?:\.\d+)+)', full_rek_raw)
                clean_rek = match_rek.group(1) if match_rek else full_rek_raw 

                full_kgt_raw = "".join(kgt_raw).replace(" ", "")
                match_kgt = re.search(r'(\d{2}\.\d{2}\.\d{2}\.?)', full_kgt_raw)
                clean_kgt = match_kgt.group(1) if match_kgt else full_kgt_raw

                nama_kegiatan = ""
                if clean_kgt:
                    search_kgt = clean_kgt.rstrip('.')
                    nama_kegiatan = await get_nama_kegiatan_from_kode(search_kgt, db)

                transaksi_list.append(
                    TransaksiItem(
                        tanggal=tgl or "",
                        kode_kegiatan=clean_kgt or "",
                        kode_rekening=clean_rek or "", 
                        no_bukti=bukti or "",
                        uraian=uraian_final,
                        nama_kegiatan=nama_kegiatan,
                        penerimaan=masuk,
                        pengeluaran=keluar,
                        saldo=sld,
                    )
                )

        if IS_DEV:
            print_debug_table(transaksi_list)

        return BKUData(transaksi=transaksi_list)

    finally:
        doc.close()

@api_router.post("/extract-bku")
async def extract_bku(
    files: List[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
):

    all_transaksi = []

    for idx, file in enumerate(files, start=1):
        try:
            if not file.filename.lower().endswith(".pdf"):
                raise HTTPException(
                    status_code=400,
                    detail=f"File bukan PDF: {file.filename}"
                )

            pdf_bytes = await file.read()

            if not pdf_bytes:
                raise ValueError("File PDF kosong")


            bku = await extract_bku_data(pdf_bytes, db)

            if not bku or not hasattr(bku, "transaksi"):
                raise ValueError("Hasil extract BKU tidak valid")

            all_transaksi.extend(bku.transaksi)

        except HTTPException:
            raise  # biarkan FastAPI tangani

        except Exception as e:
            logger.exception(
                f"❌ Gagal memproses file {file.filename}"
            )
            raise HTTPException(
                status_code=500,
                detail=f"Gagal memproses file {file.filename}: {str(e)}"
            )

    # logger.info(f"TOTAL transaksi gabungan: {len(all_transaksi)}")

    return {
        "status": "success",
        "data": {
            "transaksi": all_transaksi
        }
    }


# -----------------------------
# BHP/BHM extraction and apply
# -----------------------------
class BHPItem(BaseModel):
    tanggal: Optional[str] = ""
    kode_kegiatan: Optional[str] = ""
    kode_rekening: Optional[str] = ""
    no_bukti: Optional[str] = ""
    id_barang: Optional[str] = ""
    uraian: Optional[str] = ""
    jumlah_barang: Optional[float] = 0
    harga_satuan: Optional[float] = 0
    realisasi: Optional[float] = 0


def normalize_no_bukti_backend(val: str) -> str:
    if not val: return ''
    return re.sub(r'\W', '', val).strip().lower()


def normalize_date_backend(val: str) -> str:
    # Attempt to parse many formats and return dd-mm-yyyy
    if not val: return ''
    try:
        d = parse_date(val, dayfirst=True)
        return d.strftime('%d-%m-%Y')
    except Exception:
        # fallback: return original if looks like dd-mm-yyyy
        if re.match(r'\d{2}[-/]\d{2}[-/]\d{4}', val):
            return val.replace('/', '-')
        return ''

async def extract_bhp_data(pdf_content: bytes, filename: Optional[str] = None):
    """Extract simple BHP/BHM table rows from PDF pages. Returns list of dicts.
    Adds detailed logging for debugging: prints per-file, per-page token x,y,text and a final summary."""
    doc = fitz.open(stream=pdf_content, filetype='pdf')
    results = []

    try:
 
        for page_num, page in enumerate(doc, start=1):
            words = page.get_text('words')
            lines = {}
            for w in words:
                y = round(w[1])
                found = False
                for existing in list(lines.keys()):
                    if abs(existing - y) <= 8:
                        lines[existing].append(w)
                        found = True
                        break
                if not found:
                    lines[y] = [w]

            for y in sorted(lines.keys()):
                row_words = sorted(lines[y], key=lambda x: x[0])
                # build token objects with x,y,text and cleaned text for robust matching
                token_objs = []
                for w in row_words:
                    text = w[4].strip()
                    if not text:
                        continue
                    token_objs.append({
                        'x': round(w[0]),
                        'y': round(w[1]),
                        'text': text,
                        'clean': text.rstrip('.,')
                    })
                if not token_objs:
                    continue
                # prepare lists
                tokens = [t['text'] for t in token_objs]
                clean_tokens = [t['clean'] for t in token_objs]

                # Try to find date token using cleaned tokens
                date_token = None
                date_idx = None
                for i, tok in enumerate(clean_tokens):
                    if re.match(r'\d{2}[-/]\d{2}[-/]\d{4}', tok) or re.match(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', tok):
                        date_token = tokens[i]  # original form
                        date_idx = i
                        break

                # require date + at least 5 tokens to consider as table row
                if date_token is None or len(clean_tokens) - (date_idx + 1) < 4:
                    continue

                # pointer after date
                ptr = date_idx + 1

                # prepare fields and location maps
                kode_kegiatan = ''
                kode_rekening = ''
                no_bukti = ''
                id_barang = ''
                jumlah_barang = 0
                harga_satuan = 0
                realisasi = 0
                uraian = ''
                field_locs: Dict[str, List[dict]] = {}

                # find kode_kegiatan: first dotted token after date
                for i in range(ptr, len(clean_tokens)):
                    tok = clean_tokens[i]
                    if re.match(r'^\d+(?:\.\d+)+$', tok):
                        raw_kode = tokens[i].strip().rstrip('.')
                        # ensure trailing dot so it matches entries in master_kegiatan like '03.03.19.'
                        kode_kegiatan = raw_kode + '.' if raw_kode else ''
                        field_locs['kode_kegiatan'] = [{'x': token_objs[i]['x'], 'y': token_objs[i]['y'], 'text': token_objs[i]['text']}]
                        ptr = i+1
                        break

                # find kode_rekening: next dotted token
                for i in range(ptr, len(clean_tokens)):
                    tok = clean_tokens[i]
                    if re.match(r'^\d+(?:\.\d+)+$', tok):
                        kode_rekening = tokens[i].rstrip('.')
                        field_locs['kode_rekening'] = [{'x': token_objs[i]['x'], 'y': token_objs[i]['y'], 'text': token_objs[i]['text']}]
                        ptr = i+1
                        break

                # next possible no_bukti (should be non-dotted / alphanumeric like BPU01)
                if ptr < len(clean_tokens):
                    no_bukti = tokens[ptr]
                    field_locs['no_bukti'] = [{'x': token_objs[ptr]['x'], 'y': token_objs[ptr]['y'], 'text': token_objs[ptr]['text']}]
                    ptr += 1

                # next possible id_barang
                if ptr < len(clean_tokens) and re.match(r'^\d+(?:\.\d+)+$', clean_tokens[ptr]):
                    id_barang = tokens[ptr].rstrip('.')
                    field_locs['id_barang'] = [{'x': token_objs[ptr]['x'], 'y': token_objs[ptr]['y'], 'text': token_objs[ptr]['text']}]
                    ptr += 1

                # capture numeric tokens at end (scan backward to find up to 3 numbers)
                num_indices = []
                for i in range(len(clean_tokens)-1, ptr-1, -1):
                    if re.search(r'\d', clean_tokens[i]):
                        num_indices.append(i)
                        if len(num_indices) >= 3:
                            break
                num_indices = list(reversed(num_indices))
                # assign numeric values if present
                if len(num_indices) >= 1:
                    try:
                        # map last three numbers to jumlah, harga, realisasi where possible
                        if len(num_indices) >= 3:
                            j_idx, h_idx, r_idx = num_indices[0], num_indices[1], num_indices[2]
                            jumlah_text = token_objs[j_idx]['text']
                            harga_text = token_objs[h_idx]['text']
                            real_text = token_objs[r_idx]['text']
                            jumlah_barang = float(re.sub(r'[^0-9\-,.]','', jumlah_text).replace('.', '').replace(',', '.'))
                            harga_satuan = float(re.sub(r'[^0-9\-,.]','', harga_text).replace('.', '').replace(',', '.'))
                            realisasi = float(re.sub(r'[^0-9\-,.]','', real_text).replace('.', '').replace(',', '.'))
                            field_locs['jumlah'] = [{'x': token_objs[j_idx]['x'], 'y': token_objs[j_idx]['y'], 'text': jumlah_text}]
                            field_locs['harga'] = [{'x': token_objs[h_idx]['x'], 'y': token_objs[h_idx]['y'], 'text': harga_text}]
                            field_locs['realisasi'] = [{'x': token_objs[r_idx]['x'], 'y': token_objs[r_idx]['y'], 'text': real_text}]
                        elif len(num_indices) == 2:
                            h_idx, r_idx = num_indices[0], num_indices[1]
                            harga_text = token_objs[h_idx]['text']
                            real_text = token_objs[r_idx]['text']
                            harga_satuan = float(re.sub(r'[^0-9\-,.]','', harga_text).replace('.', '').replace(',', '.'))
                            realisasi = float(re.sub(r'[^0-9\-,.]','', real_text).replace('.', '').replace(',', '.'))
                            field_locs['harga'] = [{'x': token_objs[h_idx]['x'], 'y': token_objs[h_idx]['y'], 'text': harga_text}]
                            field_locs['realisasi'] = [{'x': token_objs[r_idx]['x'], 'y': token_objs[r_idx]['y'], 'text': real_text}]
                    except Exception:
                        pass
                uraian_parts = []
                for i in range(ptr, len(token_objs)):
                    if i in num_indices:
                        break
                    text = token_objs[i]['text']
                    # include tokens that have alphabetic chars
                    if re.search(r'[A-Za-z]', text):
                        uraian_parts.append(text)
                uraian = ' '.join(uraian_parts).strip()
                if uraian_parts:
                    field_locs['uraian'] = [{'x': token_objs[i]['x'], 'y': token_objs[i]['y'], 'text': token_objs[i]['text']} for i in range(ptr, ptr+len(uraian_parts))]
                results.append({
                    'tanggal': date_token,
                    'kode_kegiatan': kode_kegiatan,
                    'kode_rekening': kode_rekening,
                    'no_bukti': no_bukti,
                    'id_barang': id_barang,
                    'uraian': uraian,
                    'jumlah_barang': jumlah_barang,
                    'harga_satuan': harga_satuan,
                    'realisasi': realisasi,
                    'page': page_num,
                    'y': y,
                    'locations': field_locs,
                })
    finally:
        doc.close()

    return results

@api_router.post('/extract-bhp')
async def extract_bhp(
    files: List[UploadFile] = File(...),
    jenis: str = Form('BHP')
):
    """Extract BHP/BHM files and return list of parsed rows."""
    if len(files) > 24:
        raise HTTPException(status_code=400, detail='Maksimal 24 file diperbolehkan')

    all_rows = []
    for idx, f in enumerate(files, start=1):
        if not f.filename.lower().endswith('.pdf'):
            raise HTTPException(status_code=400, detail=f'File bukan PDF: {f.filename}')
        pdf = await f.read()
        rows = await extract_bhp_data(pdf, filename=f.filename)
        # annotate rows with filename for combined summary
        for r in rows:
            r['filename'] = f.filename
        # logger.info(f"Extract endpoint: file {f.filename} parsed {len(rows)} rows")
        all_rows.extend(rows)
    return {'status': 'success', 'data': all_rows}


@api_router.post('/apply-bhp')
async def apply_bhp(
    items: List[BHPItem],
    preview: bool = Query(False, description="If true, do not modify DB; return planned deletions/insertions"),
    db: AsyncSession = Depends(get_db)
):
    """Apply BHP/BHM records: for each no_bukti group, detect existing transaksi with same normalized no_bukti.
    If `preview=True` the endpoint returns planned deletions and insert counts without changing the DB.
    If `preview=False` (default) existing matching rows are deleted and replaced with extracted rows.
    """
    stmt = select(Transaksi) 
    res = await db.execute(stmt)
    transaksi_rows = res.scalars().all()

    # Group extracted items by normalized no_bukti
    groups = {}
    for b in items:
        nb = normalize_no_bukti_backend(b.no_bukti or '')
        if not nb:
            continue
        groups.setdefault(nb, []).append(b)
    group_rows = []
    planned_changes = {}
    for nb, rows in groups.items():
        # Filtering in Python (Memory) is slower than SQL filtering
        matches = [t for t in transaksi_rows if normalize_no_bukti_backend(t.no_bukti) == nb]
        
        # group_rows.append([nb, len(rows), len(matches)]) # DI-COMMENT (Bagian logging)
        
        # prepare planned delete details for preview
        planned_changes[nb] = {
            'extracted_count': len(rows),
            'existing_matches': [
                {
                    'id': t.id,
                    'tanggal': t.tanggal,
                    'no_bukti': t.no_bukti,
                    'uraian': t.uraian,
                    'volume': t.volume,
                    'harga_satuan': t.harga_satuan,
                    'pengeluaran': t.pengeluaran
                }
                for t in matches
            ]
        }
    if preview:
        total_planned_deletes = sum(len(v['existing_matches']) for v in planned_changes.values())
        total_planned_inserts = sum(v['extracted_count'] for v in planned_changes.values())
        return {
            'status': 'preview',
            'groups': planned_changes,
            'planned_deletes': total_planned_deletes,
            'planned_inserts': total_planned_inserts
        }

    # Otherwise perform delete+insert and report details
    modified_insert = 0
    modified_delete = 0
    deleted_ids: List[int] = []
    deleted_rows_details: List[dict] = []
    inserted_rows: List[dict] = []

    try:
        for nb, rows in groups.items():
            # find existing transaksi matching this no_bukti
            matches = [t for t in transaksi_rows if normalize_no_bukti_backend(t.no_bukti) == nb]

            # delete existing matches and record their details
            for t in matches:
                deleted_ids.append(t.id if getattr(t, 'id', None) is not None else None)
                deleted_rows_details.append({
                    'id': t.id,
                    'tanggal': t.tanggal,
                    'no_bukti': t.no_bukti,
                    'uraian': t.uraian,
                    'volume': t.volume,
                    'harga_satuan': t.harga_satuan,
                    'pengeluaran': t.pengeluaran
                })
                await db.delete(t)
                modified_delete += 1

            # insert new transaksi rows based on BHP items for this no_bukti
            for b in rows:
                tanggal = normalize_date_backend(b.tanggal or '') or ''
                kode_kegiatan = b.kode_kegiatan or ''
                kode_rekening = b.kode_rekening or ''
                no_bukti = b.no_bukti or ''
                uraian = b.uraian or ''

                vol = float(b.jumlah_barang or 0) if b.jumlah_barang and b.jumlah_barang > 0 else 1.0
                harga = 0.0
                if b.harga_satuan and b.harga_satuan > 0:
                    harga = float(b.harga_satuan)
                elif b.realisasi and vol > 0:
                    harga = float(b.realisasi) / vol

                pengeluaran = float(b.realisasi) if b.realisasi and b.realisasi > 0 else (vol * harga)

                new_t = Transaksi(
                    tanggal=tanggal,
                    kode_kegiatan=kode_kegiatan,
                    kode_rekening=kode_rekening,
                    no_bukti=no_bukti,
                    uraian=uraian,
                    volume=vol,
                    satuan='',
                    harga_satuan=harga,
                    penerimaan=0,
                    pengeluaran=pengeluaran,
                    saldo=0
                )
                db.add(new_t)
                await db.flush() 
                
                inserted_rows.append({
                    'id': getattr(new_t, 'id', None),
                    'tanggal': new_t.tanggal,
                    'no_bukti': new_t.no_bukti,
                    'uraian': new_t.uraian,
                    'volume': new_t.volume,
                    'harga_satuan': new_t.harga_satuan,
                    'pengeluaran': new_t.pengeluaran,
                })
                modified_insert += 1

        await db.commit()            
    except Exception as e:
        await db.rollback()
        logger.error(f"apply_bhp failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f'Gagal apply BHP: {str(e)}')

    return {
        'status': 'success',
        'deleted': modified_delete,
        'inserted': modified_insert,
        'deleted_ids': deleted_ids,
        'deleted_rows': deleted_rows_details,
        'inserted_rows': inserted_rows
    }

@api_router.post("/save-transaksi")
async def save_transaksi(
    data: BKUData,
    db: AsyncSession = Depends(get_db),
    force: bool = Query(False, description="Set true to actually perform the delete+insert. Without it the endpoint returns a preview.")
):
    """
    Save transaksi data: Mengosongkan tabel transaksi and re-insert payload.
    By default (force=False) returns a preview with counts of rows that would be deleted/inserted.
    Set force=true to actually commit the changes.
    """
    try:
        # Fetch current transaksi for preview computation
        stmt = select(Transaksi)
        res = await db.execute(stmt)
        current_rows = res.scalars().all()
        total_current = len(current_rows)

        # Build sets of normalized no_bukti values from current DB and payload
        existing_no_bukti = set()
        for t in current_rows:
            if t.no_bukti:
                existing_no_bukti.add(normalize_no_bukti_backend(t.no_bukti))

        payload_no_bukti = set()
        for t in data.transaksi:
            if getattr(t, 'no_bukti', None):
                payload_no_bukti.add(normalize_no_bukti_backend(t.no_bukti))

        # no_bukti values that would be removed (present in DB but NOT in payload)
        to_remove_no_bukti = set(x for x in existing_no_bukti if x and x not in payload_no_bukti)

        # count rows that would be removed (by matching normalized no_bukti)
        remove_rows_count = sum(1 for t in current_rows if t.no_bukti and normalize_no_bukti_backend(t.no_bukti) in to_remove_no_bukti)

        # specifically count BPU-like rows among those (BHP/BHM typical prefix)
        remove_bpu_count = sum(1 for t in current_rows if t.no_bukti and normalize_no_bukti_backend(t.no_bukti).startswith('bpu') and normalize_no_bukti_backend(t.no_bukti) in to_remove_no_bukti)

        # If not forced, return preview information instead of committing
        if not force:
            logger.info(f"save_transaksi preview: total_current={total_current}, would_delete_rows={remove_rows_count}, would_delete_bpu={remove_bpu_count}, would_insert={len(data.transaksi)}")
            return {
                "status": "preview",
                "total_current": total_current,
                "would_delete_rows": remove_rows_count,
                "would_delete_bpu_rows": remove_bpu_count,
                "would_delete_no_bukti": list(to_remove_no_bukti)[:200],
                "would_insert": len(data.transaksi)
            }

        await db.execute(delete(Transaksi))

        # 2. INSERT payload
        saved_count = 0
        for t in data.transaksi:
            # Ambil nama kegiatan dari master jika tidak ada di payload
            nama_kegiatan = t.nama_kegiatan
            if not nama_kegiatan and t.kode_kegiatan:
                nama_kegiatan = await get_nama_kegiatan_from_kode(t.kode_kegiatan, db)
            
            new_item = Transaksi(
                tanggal=t.tanggal,
                kode_kegiatan=t.kode_kegiatan or "",
                kode_rekening=t.kode_rekening or "",
                no_bukti=t.no_bukti,
                uraian=t.uraian,
                # preserve volume/satuan/harga_satuan if provided by payload
                volume=float(t.volume or 0),
                satuan=t.satuan or '',
                harga_satuan=float(t.harga_satuan or 0.0),
                penerimaan=t.penerimaan,
                pengeluaran=t.pengeluaran,
                saldo=t.saldo
            )
            db.add(new_item)
            saved_count += 1
        
        # Commit once
        await db.commit()
        logger.info(f"Regenerate Berhasil: {saved_count} transaksi baru disimpan.")
        return {
            "status": "success", 
            "message": "Data lama telah dihapus dan diganti dengan data baru",
            "saved_count": saved_count
        }

    except Exception as e:
        await db.rollback()
        logger.error(f"Error saat regenerate transaksi: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Gagal memproses data: {str(e)}")

def terbilang(n: float) -> str:
    satuan = ["", "satu", "dua", "tiga", "empat", "lima", "enam", "tujuh", "delapan", "sembilan", "sepuluh", "sebelas"]
    n = int(n)
    if n < 0:
        return "minus " + terbilang(-n)
    if n == 0:
        return "" 
    if n < 12:
        return satuan[n]
    if n < 20:
        return terbilang(n - 10) + " belas"
    if n < 100:
        hasil = terbilang(n // 10) + " puluh"
        sisa = terbilang(n % 10)
        return f"{hasil} {sisa}".strip()
    if n < 200:
        return ("seratus " + terbilang(n - 100)).strip()
    if n < 1000:
        hasil = terbilang(n // 100) + " ratus"
        sisa = terbilang(n % 100)
        return f"{hasil} {sisa}".strip()
    if n < 2000:
        return ("seribu " + terbilang(n - 1000)).strip()
    if n < 1000000:
        hasil = terbilang(n // 1000) + " ribu"
        sisa = terbilang(n % 1000)
        return f"{hasil} {sisa}".strip()
    if n < 1000000000:
        hasil = terbilang(n // 1000000) + " juta"
        sisa = terbilang(n % 1000000)
        return f"{hasil} {sisa}".strip()
    return "angka terlalu besar"

def get_thp_from_bulan(bulan: str) -> str:
    """Determine THP based on month (1-6 = THP I, 7-12 = THP II)"""
    try:
        bulan_num = int(bulan)
        if 1 <= bulan_num <= 6:
            return "THP I"
        else:
            return "THP II"
    except:
        return "THP I"

def generate_nomor_kwitansi(no_bukti: str, kode_kegiatan: str, bulan: str, tahun: str) -> str:
    """Generate kwitansi number with format: no_bukti/kode_kegiatan/THP/tahun"""
    thp = get_thp_from_bulan(bulan)
    return f"{no_bukti}/{kode_kegiatan}/{thp}/{tahun}"

def draw_hatching_pattern(c, x, y, width, height, density=3):
    """
    Membuat efek arsir (hatching) silang untuk background.
    """
    c.saveState()
    path = c.beginPath()
    path.rect(x, y, width, height)
    c.clipPath(path, stroke=0)
    
    c.setStrokeColor(colors.grey)
    c.setLineWidth(0.3)
    for i in range(-int(height), int(width), density):
        c.line(x + i, y, x + i + height, y + height)
    for i in range(0, int(width + height), density):
        c.line(x + i, y + height, x + i + height, y)
        
    c.restoreState()

def draw_wrapped_text_with_lines(c, text, x_label, x_colon, x_text, x_end, y, font_name, font_size, label_text=""):
    """
    Fungsi helper untuk menulis label, titik dua, dan teks isian yang bisa wrap (turun baris).
    Otomatis menggambar garis bawah di setiap baris isian.
    """
    # 1. Gambar Label dan Titik Dua (hanya di baris pertama)
    c.setFont(font_name, font_size)
    if label_text:
        c.drawString(x_label, y, label_text)
        c.drawString(x_colon, y, ":")

    # 2. Hitung lebar area teks yang tersedia
    available_width = x_end - x_text
    
    # 3. Pecah teks menjadi beberapa baris jika panjang
    lines = simpleSplit(text, font_name, font_size, available_width)
    
    # Jika teks kosong, setidaknya buat satu garis kosong
    if not lines:
        lines = [""]

    line_height = font_size + 4  # Jarak antar baris
    
    for i, line in enumerate(lines):
        # Tulis teks
        c.drawString(x_text, y, line)
        
        # Gambar garis bawah (Underline) full sampai ujung kanan
        c.setLineWidth(0.5) # Garis tipis untuk isian
        c.line(x_text, y - 2, x_end, y - 2)
        
        # Kurangi Y untuk baris berikutnya
        y -= line_height
        
    return y # Kembalikan posisi Y terakhir agar elemen berikutnya menyesuaikan

def generate_kwitansi_pdf_v2(kwitansi, pengaturan):
    buffer = BytesIO()
    
    # --- KONFIGURASI HALAMAN ---
    F4_WIDTH = 21.5 * cm
    F4_HEIGHT = 33.0 * cm
    # Canvas portrait, tapi kita menggambar di area tertentu (simulasi landscape)
    c = canvas.Canvas(buffer, pagesize=(F4_WIDTH, F4_HEIGHT))

    # --- SETTING LAYOUT AREA ---
    margin_x = 1.0 * cm
    top_y = F4_HEIGHT - 2.0 * cm
    
    kwitansi_width = F4_WIDTH - (2 * margin_x)
    # Tinggi area dinamis, tapi kita set base area dulu
    kwitansi_height = 14.0 * cm 
    bottom_y = top_y - kwitansi_height

    # --- 1. HEADER & JUDUL ---
    header_height = 1.2 * cm
    header_y = top_y - 2.0 * cm 
    header_margin_x = margin_x + 1.0 * cm
    header_width = kwitansi_width - 2.0 * cm
    
    # Arsir Background Header
    c.setLineWidth(1)
    draw_hatching_pattern(c, header_margin_x, header_y, header_width, header_height, density=4)
    c.setStrokeColor(colors.black)
    c.rect(header_margin_x, header_y, header_width, header_height)
    
    # Teks Judul
    c.setFont("Times-Bold", 20)
    c.setFillColor(colors.black)
    c.drawCentredString(F4_WIDTH / 2, header_y + 0.35 * cm, "KWITANSI")

    # --- 2. ISI FORMULIR ---
    content_start_y = header_y - 1.5 * cm
    
    # Koordinat Kolom
    left_col_x = margin_x + 1.0 * cm     # Label
    colon_x = margin_x + 4.5 * cm        # Titik dua
    value_x = margin_x + 4.8 * cm        # Awal isian
    line_end_x = margin_x + kwitansi_width - 1.0 * cm # Ujung kanan garis
    
    c.setFont("Times-Roman", 12)
    current_y = content_start_y
    line_gap = 0.8 * cm # Jarak antar field utama

    # A. Nomor
    # Kita pakai wrap function juga biar konsisten style garisnya, meski biasanya nomor pendek
    current_y = draw_wrapped_text_with_lines(
        c, kwitansi.nomor_kwitansi, 
        left_col_x, colon_x, value_x, value_x + 6*cm, # Garis nomor tidak perlu sampai ujung
        current_y, "Times-Roman", 12, "Nomor"
    )
    current_y -= line_gap

    # B. Telah Terima Dari
    current_y = draw_wrapped_text_with_lines(
        c, f"BENDAHARA BOSP {pengaturan.nama_sekolah}", 
        left_col_x, colon_x, value_x, line_end_x,
        current_y, "Times-Roman", 12, "Telah terima dari"
    )
    current_y -= line_gap

    # C. Uang Sejumlah (Terbilang) - Font Bold
    # Convert jumlah ke terbilang (pastikan fungsi terbilang ada)
    try:
        txt_terbilang = f"{terbilang(kwitansi.jumlah).capitalize()} rupiah"
    except:
        txt_terbilang = "................ rupiah" # Fallback

    current_y = draw_wrapped_text_with_lines(
        c, txt_terbilang, 
        left_col_x, colon_x, value_x, line_end_x,
        current_y, "Times-Roman", 12, "Uang sejumlah"
    )
    current_y -= line_gap

    # D. Untuk Pembayaran (WRAP TEXT PENTING DISINI)
    # Ini akan otomatis turun baris dan mendorong elemen di bawahnya
    current_y = draw_wrapped_text_with_lines(
        c, kwitansi.nama_kegiatan, 
        left_col_x, colon_x, value_x, line_end_x,
        current_y, "Times-Roman", 12, "Untuk pembayaran"
    )
    sep_y = current_y - 0.5 * cm
    c.setLineWidth(0.5)
    # c.line(left_col_x, sep_y, line_end_x, sep_y) # (Opsional: garis pemisah form dan ttd)

    # --- 4. TANDA TANGAN ---
    ttd_base_y = sep_y - 1.0 * cm
    
    # Koordinat X Tanda Tangan
    col_1_x = margin_x + 1.0 * cm                  # Kiri (KS)
    col_2_x = margin_x + (kwitansi_width / 2) - 2*cm  # Tengah (Bendahara)
    col_3_x = margin_x + kwitansi_width - 5.5 * cm    # Kanan (Penerima)

    c.setFont("Times-Roman", 11)
    
    # Header TTD
    c.drawString(col_1_x, ttd_base_y, "Setuju Dibayar")
    c.drawString(col_1_x, ttd_base_y - 0.5*cm, "Kepala Sekolah")
    
    c.drawString(col_2_x, ttd_base_y, f"Lunas Dibayar {kwitansi.tanggal}") 
    c.drawString(col_2_x, ttd_base_y - 0.5*cm, "Bendahara BOSP")

    kota_ttd = getattr(pengaturan, 'tempat_surat', None) or 'Jember'
    c.drawString(col_3_x, ttd_base_y, f"{kota_ttd}, {kwitansi.tanggal}")
    c.drawString(col_3_x, ttd_base_y - 0.5*cm, "Penerima")

    # Nama & NIP (Jarak sekitar 2.5cm - 3cm ke bawah)
    name_y = ttd_base_y - 3.0 * cm
    nip_y = name_y - 0.5 * cm

    # TTD Kiri (KS)
    c.setFont("Times-Bold", 11)
    c.drawString(col_1_x, name_y, pengaturan.nama_kepala_sekolah)
    c.setFont("Times-Roman", 11)
    c.drawString(col_1_x, nip_y, f"NIP. {pengaturan.nip_kepala_sekolah}")

    # TTD Tengah (Bendahara)
    c.setFont("Times-Bold", 11)
    c.drawString(col_2_x, name_y, pengaturan.nama_bendahara)
    c.setFont("Times-Roman", 11)
    c.drawString(col_2_x, nip_y, f"NIP. {pengaturan.nip_bendahara}")

    # TTD Kanan (Penerima)
    c.line(col_3_x, name_y, col_3_x + 4.5*cm, name_y) # Garis tanda tangan

    # --- 5. KOTAK NOMINAL (Rp.) ---
    # Posisi di bawah NIP KS
    rp_box_y = nip_y - 1.5 * cm 
    rp_box_x = col_1_x + 1.0 * cm 
    rp_box_w = 6.0 * cm
    rp_box_h = 1.0 * cm

    # Arsir Kotak Rp
    draw_hatching_pattern(c, rp_box_x, rp_box_y, rp_box_w, rp_box_h, density=3)
    c.setStrokeColor(colors.black)
    c.setLineWidth(1)
    c.rect(rp_box_x, rp_box_y, rp_box_w, rp_box_h)

    # Label Rp. (Di luar kotak sebelah kiri)
    c.setFont("Times-Bold", 14)
    c.drawString(rp_box_x - 1.0 * cm, rp_box_y + 0.3*cm, "Rp.")
    
    # Angka Nominal (Di dalam kotak)
    formatted_jumlah = f"{kwitansi.jumlah:,.00f}".replace(",", "X").replace(".", ",").replace("X", ".")
    center_x_box = rp_box_x + (rp_box_w / 2)
    c.drawCentredString(center_x_box, rp_box_y + 0.3*cm, formatted_jumlah)

    # --- DRAW BORDER LUAR (FINAL) ---
    # Gambar border luar berdasarkan posisi paling bawah konten agar dinamis
    # Tapi untuk menjaga tampilan kartu, kita pakai fix height minimal
    final_bottom_y = rp_box_y - 1.0 * cm
    
    # Jika konten meluber melebihi area standar kwitansi, sesuaikan border
    if final_bottom_y < bottom_y:
        actual_bottom_y = final_bottom_y
    else:
        actual_bottom_y = bottom_y
        
    actual_height = top_y - actual_bottom_y

    # Garis Tebal Luar
    c.setLineWidth(1.5)
    c.rect(margin_x, actual_bottom_y, kwitansi_width, actual_height)
    
    # Garis Tipis Dalam
    c.setLineWidth(0.5)
    gap = 0.15 * cm
    c.rect(margin_x + gap, actual_bottom_y + gap, kwitansi_width - 2*gap, actual_height - 2*gap)

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer.getvalue()

@api_router.get("/kwitansi/{id}/pdf")
async def get_kwitansi_pdf(
    id: int,
    db: AsyncSession = Depends(get_db)
):
    # 1. Ambil Data Kwitansi
    stmt = select(Kwitansi).where(Kwitansi.id == id)
    result = await db.execute(stmt)
    kwitansi = result.scalar_one_or_none()
    
    if not kwitansi:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # 2. Ambil Pengaturan
    stmt2 = select(Pengaturan).limit(1)
    result2 = await db.execute(stmt2)
    pengaturan = result2.scalar_one_or_none()
    
    if not pengaturan:
        raise HTTPException(status_code=404, detail="Pengaturan belum diisi")

    # 3. Generate Bagian 1: KWITANSI (Halaman Utama)
    try:
        kw_pdf_bytes = generate_kwitansi_pdf_v2(kwitansi, pengaturan)
    except Exception as e:
        logger.error(f"Gagal generate Kwitansi PDF: {e}")
        raise HTTPException(status_code=500, detail="Gagal membuat PDF Kwitansi")

    # 4. Generate Bagian 2: BAST & FOTO (Jika ada no_bukti)
    bast_pdf_bytes = None
    if kwitansi.no_bukti:
        # Ambil item transaksi dari database untuk tabel barang di BAST
        stmt_trx = select(Transaksi).where(Transaksi.no_bukti == kwitansi.no_bukti).order_by(Transaksi.id)
        res_trx = await db.execute(stmt_trx)
        transaksi_list = res_trx.scalars().all()

        if transaksi_list:
            try:
                # Panggil fungsi BAST yang sudah include FOTO di dalamnya
                bast_pdf_bytes = generate_bast_pdf(kwitansi, pengaturan, transaksi_list)
            except Exception as e:
                # Jika BAST/Foto gagal, kita log saja agar kwitansi tetap bisa terdownload
                logger.error(f"Gagal generate BAST/Foto PDF: {e}")

    # 5. Proses Penggabungan (Merging)
    final_buffer = BytesIO()
    
    if bast_pdf_bytes:
        # Gabungkan Kwitansi + BAST + Foto
        merger = PdfMerger()
        merger.append(BytesIO(kw_pdf_bytes))
        merger.append(BytesIO(bast_pdf_bytes))
        merger.write(final_buffer)
        merger.close()
    else:
        # Jika tidak ada BAST atau no_bukti kosong, kirim Kwitansi saja
        final_buffer.write(kw_pdf_bytes)

    final_buffer.seek(0)
    filename = f"SPJ_Lengkap_{kwitansi.no_bukti or kwitansi.id}.pdf"

    return StreamingResponse(
        final_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

@api_router.get("/transaksi")
async def get_transaksi(
    db: AsyncSession = Depends(get_db),
    no_bukti: Optional[str] = None
):
    try:
        if no_bukti:
            stmt = select(Transaksi).where(Transaksi.no_bukti == no_bukti)
        else:
            # FILTER & URUTKAN
            stmt = select(Transaksi).where(
                and_(
                    Transaksi.no_bukti.isnot(None),
                    Transaksi.no_bukti != "",
                    Transaksi.no_bukti.like("BPU%")
                )
            ).order_by(
                # 1. Urutkan berdasarkan jumlah karakter (agar BPU1-9 di atas BPU10-99)
                func.length(Transaksi.no_bukti).asc(), 
                # 2. Urutkan secara alfabetis untuk nomor dengan jumlah karakter yang sama
                Transaksi.no_bukti.asc()
            )
        
        result = await db.execute(stmt)
        transaksi_list = result.scalars().all()
        
        return {"status": "success", "data": transaksi_list}
    except Exception as e:
        logger.error(f"Error getting transaksi: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/transaksi/sumif/{no_bukti}")
async def get_sumif_no_bukti(
    no_bukti: str,
    db: AsyncSession = Depends(get_db)
):
    """Get SUMIF total for no_bukti (like Excel SUMIF)"""
    try:
        total = await calculate_sumif_no_bukti(no_bukti, db)
        return {"status": "success", "no_bukti": no_bukti, "total": total}
    except Exception as e:
        logger.error(f"Error calculating SUMIF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error calculating SUMIF: {str(e)}")


# Helper: calculate SUM of pengeluaran for a given no_bukti (Excel-like SUMIF)
async def calculate_sumif_no_bukti(no_bukti: str, db: AsyncSession) -> float:
    try:
        stmt = select(func.sum(Transaksi.pengeluaran)).where(Transaksi.no_bukti == no_bukti)
        result = await db.execute(stmt)
        total = result.scalar() or 0.0
        return float(total)
    except Exception as e:
        logger.error(f"Error in calculate_sumif_no_bukti: {str(e)}")
        return 0.0


# Pastikan import ini ada
# Pastikan Anda mengimport Kwitansi dari models/file tempat class didefinisikan
# from models import Transaksi, Kwitansi 

@api_router.delete("/transaksi/{no_bukti}")
async def delete_transaksi(no_bukti: str, db: AsyncSession = Depends(get_db)):
    # 1. Decode no_bukti (jaga-jaga karakter spesial)
    clean_no_bukti = unquote(no_bukti)
    
    # 2. Hapus data dari Tabel Transaksi
    stmt_transaksi = delete(Transaksi).where(Transaksi.no_bukti == clean_no_bukti)
    await db.execute(stmt_transaksi)

    # 3. Hapus data dari Tabel Kwitansi (TAMBAHAN BARU)
    # Ini akan menghapus baris di tabel kwitansi yang punya no_bukti sama
    stmt_kwitansi = delete(Kwitansi).where(Kwitansi.no_bukti == clean_no_bukti)
    await db.execute(stmt_kwitansi)

    # 4. Commit perubahan (Simpan kedua penghapusan sekaligus)
    await db.commit()
    
    return {
        "status": "success", 
        "message": f"Data Transaksi dan Kwitansi untuk no bukti '{clean_no_bukti}' berhasil dihapus"
    }

@api_router.get("/kwitansi")
async def get_kwitansi_final(db: AsyncSession = Depends(get_db)):
    """
    Menampilkan semua kwitansi lengkap, termasuk tanggal_nota, no_bast, nama_toko, npwp_toko, alamat_toko.
    """
    try:
        stmt = select(Kwitansi).order_by(Kwitansi.tanggal.asc(), Kwitansi.no_bukti.asc())
        result = await db.execute(stmt)
        kwitansi_list = result.scalars().all()

        data = []
        for k in kwitansi_list:
            data.append({
                "id": k.id,
                "nomor_kwitansi": k.nomor_kwitansi,
                "no_bukti": k.no_bukti,
                "kode_kegiatan": k.kode_kegiatan,
                "nama_kegiatan": k.nama_kegiatan,
                "tanggal": k.tanggal,
                "thp": k.thp,
                "tahun": k.tahun,
                "jumlah": k.jumlah,
                "tanggal_nota": k.tanggal_nota,
                "no_bast": k.no_bast,
                "nama_toko": k.nama_toko,
                "npwp_toko": k.npwp_toko,
                "alamat_toko": k.alamat_toko,
                "created_at": k.created_at,
                "updated_at": k.updated_at
            })

        return {"status": "success", "data": data}

    except Exception as e:
        logger.error(f"Error getting kwitansi: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/kwitansi/generate")
async def generate_kwitansi_final(db: AsyncSession = Depends(get_db)):
    """
    Generate kwitansi berdasarkan transaksi.
    Data tambahan seperti tanggal_nota, no_bast, nama_toko, npwp_toko, alamat_toko diisi kosong.
    Format tanggal otomatis dikenali.
    """
    try:
        # Hapus semua kwitansi lama
        await db.execute(delete(Kwitansi))
        await db.commit()

        # Ambil transaksi yang memiliki pengeluaran > 0
        stmt = (
            select(
                Transaksi.no_bukti,
                Transaksi.kode_kegiatan,
                MasterKegiatan.nama_kegiatan.label("nama_kegiatan_master"),
                func.sum(Transaksi.pengeluaran).label("jumlah"),
                func.min(Transaksi.tanggal).label("tanggal")
            )
            .join(
                MasterKegiatan,
                Transaksi.kode_kegiatan == MasterKegiatan.kode_kegiatan,
                isouter=True
            )
            .where(
                Transaksi.no_bukti.isnot(None),
                func.length(func.trim(Transaksi.no_bukti)) > 0,
                Transaksi.pengeluaran > 0
            )
            .group_by(
                Transaksi.no_bukti,
                Transaksi.kode_kegiatan,
                MasterKegiatan.nama_kegiatan
            )
            .order_by(
                func.min(Transaksi.tanggal).asc(),
                Transaksi.no_bukti.asc()
            )
        )

        result = await db.execute(stmt)
        rows = result.all()
        generated = 0

        for row in rows:
            if not row.tanggal or not row.jumlah:
                continue

            # Parse tanggal otomatis, format fleksibel
            try:
                tgl = parse_date(row.tanggal, dayfirst=True)
            except Exception:
                continue  # lewati jika tidak bisa diparse

            bulan = tgl.month
            tahun = tgl.year
            thp = "THP I" if bulan <= 6 else "THP II"

            nomor_kwitansi = f"{row.no_bukti}/{row.kode_kegiatan or '-'} /{thp}/{tahun}"

            kw = Kwitansi(
                nomor_kwitansi=nomor_kwitansi,
                no_bukti=row.no_bukti,
                kode_kegiatan=row.kode_kegiatan or "",
                nama_kegiatan=row.nama_kegiatan_master or "Tanpa Nama Kegiatan",
                tanggal=row.tanggal,
                thp=thp,
                tahun=str(tahun),
                jumlah=row.jumlah,
                tanggal_nota="",   # default kosong
                no_bast="",        # default kosong
                nama_toko="",      # default kosong
                npwp_toko="",      # default kosong
                alamat_toko=""     # default kosong
            )

            db.add(kw)
            generated += 1

        await db.commit()
        return {"status": "success", "generated": generated}

    except Exception as e:
        await db.rollback()
        logger.exception("ERROR GENERATE KWITANSI")
        raise HTTPException(status_code=500, detail=str(e))

# Pengaturan (Settings) Routes
@api_router.get("/pengaturan")
async def get_pengaturan(
    db: AsyncSession = Depends(get_db)
):
    """Get application settings"""
    try:
        stmt = select(Pengaturan).limit(1)
        result = await db.execute(stmt)
        pengaturan = result.scalar_one_or_none()
        
        if not pengaturan:
            return {"status": "success", "data": None}
        
        return {
            "status": "success", 
            "data": {
                "id": pengaturan.id,
                "nama_sekolah": pengaturan.nama_sekolah,
                "nama_kepala_sekolah": pengaturan.nama_kepala_sekolah,
                "nip_kepala_sekolah": pengaturan.nip_kepala_sekolah,
                "nama_bendahara": pengaturan.nama_bendahara,
                "nip_bendahara": pengaturan.nip_bendahara,
                "nama_pengurus_barang": pengaturan.nama_pengurus_barang,
                "nip_pengurus_barang": pengaturan.nip_pengurus_barang,
                "alamat_sekolah": pengaturan.alamat_sekolah,
                "tempat_surat": pengaturan.tempat_surat or ""
            }
        }
    except Exception as e:
        logger.error(f"Error getting pengaturan: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting pengaturan: {str(e)}")


@api_router.post("/pengaturan")
async def save_pengaturan(
    request: PengaturanRequest,
    db: AsyncSession = Depends(get_db)
):
    """Save or update application settings"""
    try:
        stmt = select(Pengaturan).limit(1)
        result = await db.execute(stmt)
        pengaturan = result.scalar_one_or_none()
        
        if pengaturan:
            # Update existing
            pengaturan.nama_sekolah = request.nama_sekolah
            pengaturan.nama_kepala_sekolah = request.nama_kepala_sekolah
            pengaturan.nip_kepala_sekolah = request.nip_kepala_sekolah
            pengaturan.nama_bendahara = request.nama_bendahara
            pengaturan.nip_bendahara = request.nip_bendahara
            pengaturan.nama_pengurus_barang = request.nama_pengurus_barang
            pengaturan.nip_pengurus_barang = request.nip_pengurus_barang
            pengaturan.alamat_sekolah = request.alamat_sekolah
            pengaturan.tempat_surat = request.tempat_surat
        else:
            # Create new
            pengaturan = Pengaturan(
                nama_sekolah=request.nama_sekolah,
                nama_kepala_sekolah=request.nama_kepala_sekolah,
                nip_kepala_sekolah=request.nip_kepala_sekolah,
                nama_bendahara=request.nama_bendahara,
                nip_bendahara=request.nip_bendahara,
                nama_pengurus_barang=request.nama_pengurus_barang,
                nip_pengurus_barang=request.nip_pengurus_barang,
                alamat_sekolah=request.alamat_sekolah,
                tempat_surat=request.tempat_surat
            )
            db.add(pengaturan)
        
        await db.commit()
        return {"status": "success", "message": "Pengaturan berhasil disimpan"}
    except Exception as e:
        await db.rollback()
        logger.error(f"Error saving pengaturan: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error saving pengaturan: {str(e)}")


# Master Kegiatan Routes
@api_router.get("/master-kegiatan")
async def get_master_kegiatan(
    db: AsyncSession = Depends(get_db)
):
    """Get all master kegiatan"""
    try:
        stmt = select(MasterKegiatan).order_by(MasterKegiatan.kode_kegiatan)
        result = await db.execute(stmt)
        kegiatan_list = result.scalars().all()
        
        return {"status": "success", "data": kegiatan_list}
    except Exception as e:
        logger.error(f"Error getting master kegiatan: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting master kegiatan: {str(e)}")


@api_router.post("/master-kegiatan")
async def save_master_kegiatan(
    request: MasterKegiatanRequest,
    db: AsyncSession = Depends(get_db)
):
    """Save or update master kegiatan"""
    try:
        stmt = select(MasterKegiatan).where(MasterKegiatan.kode_kegiatan == request.kode_kegiatan)
        result = await db.execute(stmt)
        kegiatan = result.scalar_one_or_none()
        
        if kegiatan:
            kegiatan.nama_kegiatan = request.nama_kegiatan
        else:
            kegiatan = MasterKegiatan(
                kode_kegiatan=request.kode_kegiatan,
                nama_kegiatan=request.nama_kegiatan
            )
            db.add(kegiatan)
        
        await db.commit()
        return {"status": "success", "message": "Master kegiatan berhasil disimpan"}
    except Exception as e:
        await db.rollback()
        logger.error(f"Error saving master kegiatan: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error saving master kegiatan: {str(e)}")

@api_router.post("/master-kegiatan/bulk")
async def save_master_kegiatan_bulk(
    request: MasterKegiatanBulkRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Bulk import master kegiatan (Replace All).
    Metode: HAPUS semua data lama, lalu INSERT data baru.
    """
    try:
        # 1. Hapus seluruh baris data di tabel MasterKegiatan
        await db.execute(delete(MasterKegiatan))
        
        # 2. Siapkan objek baru dari request
        new_data_list = []
        for item in request.data:
            kegiatan = MasterKegiatan(
                kode_kegiatan=item.kode_kegiatan,
                nama_kegiatan=item.nama_kegiatan
            )
            new_data_list.append(kegiatan)
            
        # 3. Masukkan data baru sekaligus (lebih efisien daripada add satu per satu di loop)
        if new_data_list:
            db.add_all(new_data_list)
        
        # 4. Commit transaksi (Delete dan Insert terjadi dalam satu transaksi)
        await db.commit()
        
        return {"status": "success", "saved_count": len(new_data_list)}

    except Exception as e:
        await db.rollback()
        logger.error(f"Error bulk replacing master kegiatan: {str(e)}")
        # Menangani error Foreign Key jika data ini sedang dipakai di tabel lain
        if "foreign key constraint" in str(e).lower():
             raise HTTPException(
                status_code=400, 
                detail="Gagal menghapus data lama karena data sedang digunakan di tabel lain (Foreign Key Error)."
            )
        raise HTTPException(status_code=500, detail=f"Error bulk replacing master kegiatan: {str(e)}")


from sqlalchemy import delete  # <--- Pastikan ini diimport

@api_router.post("/master-rekening-belanja/bulk")
async def save_master_rekening_belanja_bulk(
    request: MasterRekeningBelanjaBulkRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Bulk import master rekening belanja (Replace All).
    Metode: HAPUS semua data lama, lalu INSERT data baru.
    """
    try:
        # 1. Hapus seluruh data di tabel MasterRekeningBelanja
        await db.execute(delete(MasterRekeningBelanja))

        # 2. Siapkan list objek baru dari request
        new_rekening_list = []
        for item in request.data:
            rek = MasterRekeningBelanja(
                kode_rekening_belanja=item.kode_rekening_belanja,
                nama_rekening_belanja=item.nama_rekening_belanja,
                # Pastikan logic default value tetap ada seperti kode asli
                rekap_rekening_belanja=item.rekap_rekening_belanja or "",
                nilai_kapitalisasi_belanja=item.nilai_kapitalisasi_belanja or 0
            )
            new_rekening_list.append(rek)

        # 3. Masukkan data baru sekaligus (lebih efisien)
        if new_rekening_list:
            db.add_all(new_rekening_list)

        # 4. Commit transaksi (Delete dan Insert terjadi bersamaan)
        await db.commit()
        
        return {"status": "success", "saved_count": len(new_rekening_list)}

    except Exception as e:
        await db.rollback()
        logger.error(f"Error bulk replacing master rekening belanja: {str(e)}")
        
        # Penanganan khusus jika gagal hapus karena Foreign Key Constraint
        if "foreign key constraint" in str(e).lower():
             raise HTTPException(
                status_code=400, 
                detail="Gagal melakukan replace data. Data Rekening Belanja lama masih digunakan di tabel lain (transaksi/anggaran)."
            )
            
        raise HTTPException(status_code=500, detail=f"Error bulk replacing master rekening belanja: {str(e)}")


@api_router.get("/master-rekening-belanja")
async def get_master_rekening_belanja(
    db: AsyncSession = Depends(get_db)
):
    """Get all master rekening belanja"""
    try:
        stmt = select(MasterRekeningBelanja).order_by(MasterRekeningBelanja.kode_rekening_belanja)
        result = await db.execute(stmt)
        rekening_list = result.scalars().all()

        # Serialize to simple dicts for JSON response
        data = []
        for r in rekening_list:
            data.append({
                "id": r.id,
                "kode_rekening_belanja": r.kode_rekening_belanja,
                "nama_rekening_belanja": r.nama_rekening_belanja,
                "rekap_rekening_belanja": r.rekap_rekening_belanja,
                "nilai_kapitalisasi_belanja": r.nilai_kapitalisasi_belanja,
                "created_at": r.created_at,
                "updated_at": r.updated_at,
            })

        return {"status": "success", "data": data}
    except Exception as e:
        logger.error(f"Error getting master rekening belanja: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting master rekening belanja: {str(e)}")

@api_router.delete("/master-rekening-belanja/{id}")
async def delete_master_rekening(id: int, db: AsyncSession = Depends(get_db)):
    await db.execute(delete(MasterRekeningBelanja).where(MasterRekeningBelanja.id == id))
    await db.commit()
    return {"status": "success"}

@api_router.get("/rekap-transaksi")
async def get_rekap_transaksi(
    details: int = 0,
    thp: str = "all",
    db: AsyncSession = Depends(get_db)
):
    """Generate rekap:
    - Only transactions that have a matching `kode_rekening` in `master_rekening_belanja` are considered.
    - First classification is per transaction: if master.rekap_rekening_belanja is 'KIB B' or 'KIB E',
      that transaction is labeled either '<KIB> Kapitalisasi' (if harga_satuan >= nilai_kapitalisasi_belanja)
      or 'Non Kapitalisasi' otherwise.
    - All transactions are then aggregated (sum of `pengeluaran`) grouped by the final label.
    - Pass ?details=1 to include transaction-level items in each group.
    """
    try:
        # Inner join ensures only transactions with a matching master rekening are considered.
        # Exclude transactions where kode_rekening or no_bukti are null/empty
        # Build base stmt joining MasterRekening
        base_stmt = select(Transaksi, MasterRekeningBelanja).join(
            MasterRekeningBelanja,
            MasterRekeningBelanja.kode_rekening_belanja == Transaksi.kode_rekening
        ).where(Transaksi.kode_rekening != None).where(Transaksi.kode_rekening != "").where(Transaksi.no_bukti != None).where(Transaksi.no_bukti != "")

        # If thp filter is specified (1 or 2), join Kwitansi and filter by its `thp` column
        if str(thp).strip() in ("1", "2"):
            thp_map = {"1": "THP I", "2": "THP II"}
            target_thp = thp_map[str(thp).strip()]
            base_stmt = base_stmt.join(Kwitansi, Kwitansi.no_bukti == Transaksi.no_bukti).where(Kwitansi.thp == target_thp)

        result = await db.execute(base_stmt)
        rows = result.all()

        groups = {}

        for trans, master in rows:
            # base rekap from master
            base_rekap = (master.rekap_rekening_belanja or "").strip()

            # Special KIB handling: produce explicit labels per KIB type
            if base_rekap in ("KIB B", "KIB E"):
                threshold = float(master.nilai_kapitalisasi_belanja or 0)
                harga = float(trans.harga_satuan or 0)
                if harga >= threshold:
                    label = f"{base_rekap} Kapitalisasi"
                else:
                    label = f"{base_rekap} Non Kapitalisasi"
            else:
                label = base_rekap or "Unspecified"

            g = groups.get(label)
            if not g:
                g = {
                    "rekap": label,
                    "count": 0,
                    "total_pengeluaran": 0.0,
                    "kode_rekening_breakdown": {},
                    "items": []
                }
                groups[label] = g

            pel = float(trans.pengeluaran or 0)
            g["count"] += 1
            g["total_pengeluaran"] += pel

            # breakdown per kode_rekening
            kr = trans.kode_rekening or ""
            g["kode_rekening_breakdown"][kr] = g["kode_rekening_breakdown"].get(kr, 0.0) + pel

            if details:
                g["items"].append({
                    "no_bukti": trans.no_bukti,
                    "kode_rekening": trans.kode_rekening,
                    "kode_kegiatan": trans.kode_kegiatan,
                    "harga_satuan": trans.harga_satuan,
                    "pengeluaran": trans.pengeluaran
                })

        # Convert to list and sort by rekap name
        data = sorted(list(groups.values()), key=lambda x: x.get("rekap") or "")
        return {"status": "success", "data": data}

    except Exception as e:
        logger.error(f"Error generating rekap transaksi: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating rekap transaksi: {str(e)}")

def generate_placeholder_code(text: str) -> str:
    """Mengubah 'Belanja Bahan-Bahan' menjadi 'BELANJA_BAHAN_BAHAN'"""
    if not text:
        return ""
    # Ganti karakter non-alphanumeric (spasi, -, /) dengan underscore, lalu uppercase
    return re.sub(r'[^A-Z0-9]', '_', text.strip().upper())

# --- ENDPOINT: Generate Excel dari Template User ---
@api_router.post("/generate-rekap-excel")
async def generate_rekap_excel(
    file: UploadFile = File(...),
    thp: str = Form("all"),
    db: AsyncSession = Depends(get_db)
):
    try:
        # 1. === AMBIL DATA ===
        base_stmt = select(Transaksi, MasterRekeningBelanja).join(
            MasterRekeningBelanja,
            MasterRekeningBelanja.kode_rekening_belanja == Transaksi.kode_rekening
        ).where(
            Transaksi.kode_rekening != None, 
            Transaksi.kode_rekening != "", 
            Transaksi.no_bukti != None, 
            Transaksi.no_bukti != ""
        )

        if str(thp).strip() in ("1", "2"):
            thp_map = {"1": "THP I", "2": "THP II"}
            target_thp = thp_map[str(thp).strip()]
            base_stmt = base_stmt.join(Kwitansi, Kwitansi.no_bukti == Transaksi.no_bukti).where(Kwitansi.thp == target_thp)

        result = await db.execute(base_stmt)
        rows = result.all()

        # 2. === HITUNG TOTAL ===
        rekap_totals = {}
        for trans, master in rows:
            base_rekap = (master.rekap_rekening_belanja or "").strip()
            if base_rekap in ("KIB B", "KIB E"):
                threshold = float(master.nilai_kapitalisasi_belanja or 0)
                harga = float(trans.harga_satuan or 0)
                label = f"{base_rekap} Kapitalisasi" if harga >= threshold else f"{base_rekap} Non Kapitalisasi"
            else:
                label = base_rekap or "Unspecified"

            key_code = generate_placeholder_code(label)
            pel = float(trans.pengeluaran or 0)
            rekap_totals[key_code] = rekap_totals.get(key_code, 0.0) + pel

        # 3. === PROSES EXCEL ===
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell_text = cell.value.strip().upper()
                        if cell_text in rekap_totals:
                            cell.value = rekap_totals[cell_text]
                            cell.number_format = '#,##0' 

        # 4. === SIMPAN ===
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        original_name = file.filename
        
        # --- PERBAIKAN REGEX DI SINI ---
        # Menambahkan backslash sebelum tanda minus: [_\-\s]
        clean_name = re.sub(r'[_\-\s]*template[_\-\s]*', '', original_name, flags=re.IGNORECASE)

        if len(clean_name) < 5: 
            clean_name = f"Rekap_Result_{thp}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={clean_name}"}
        )

    except Exception as e:
        print(f"Error generating excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Gagal memproses Excel: {str(e)}")


@api_router.delete("/master-kegiatan/{kode_kegiatan}")
async def delete_master_kegiatan(
    kode_kegiatan: str,
    db: AsyncSession = Depends(get_db)
):
    """Delete master kegiatan"""
    try:
        stmt = select(MasterKegiatan).where(MasterKegiatan.kode_kegiatan == kode_kegiatan)
        result = await db.execute(stmt)
        kegiatan = result.scalar_one_or_none()
        
        if not kegiatan:
            raise HTTPException(status_code=404, detail="Master kegiatan tidak ditemukan")
        
        await db.delete(kegiatan)
        await db.commit()
        return {"status": "success", "message": "Master kegiatan berhasil dihapus"}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error deleting master kegiatan: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting master kegiatan: {str(e)}")

# --- GET detail kwitansi ---
@api_router.get("/kwitansi/{kwitansi_id}")
async def get_kwitansi_detail(
    kwitansi_id: int,
    db: AsyncSession = Depends(get_db)
):
    # =========================
    # 1. Ambil data Kwitansi
    # =========================
    result = await db.execute(
        select(Kwitansi).where(Kwitansi.id == kwitansi_id)
    )
    kw = result.scalar_one_or_none()

    if not kw:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # =========================
    # 2. Ambil transaksi (uraian)
    # =========================
    transaksi_list = []
    if kw.no_bukti:
        transaksi_result = await db.execute(
            select(Transaksi).where(Transaksi.no_bukti == kw.no_bukti)
        )
        transaksi_list = transaksi_result.scalars().all()

    uraian_list = [
        {
            "nama_barang": t.uraian,
            "volume": t.volume or 0,
            "satuan": t.satuan or "",
            "harga_satuan": t.harga_satuan or 0,
            "pengeluaran": t.pengeluaran or 0
        }
        for t in transaksi_list
    ]

    # =========================
    # 3. PROSES FOTO BUKTI
    # =========================
    try:
        raw_foto = json.loads(kw.foto_bukti or "[]")
    except Exception:
        raw_foto = []

    # Grouping foto berdasarkan jenis
    foto_grouped = {}

    for item in raw_foto:
        # Support data lama (string)
        if isinstance(item, str):
            jenis = "barang"
            path = item

        # Support data baru (object)
        elif isinstance(item, dict):
            jenis = item.get("jenis", "barang")
            path = item.get("path")

        else:
            continue

        if not path:
            continue

        foto_grouped.setdefault(jenis, []).append(path)

    # =========================
    # 4. RESPONSE FINAL
    # =========================
    return {
        "status": "success",
        "data": {
            "id": kw.id,
            "nomor_kwitansi": kw.nomor_kwitansi,
            "no_bukti": kw.no_bukti,
            "no_bast": kw.no_bast,
            "kode_kegiatan": kw.kode_kegiatan,
            "nama_kegiatan": kw.nama_kegiatan,
            "tanggal": kw.tanggal,
            "tanggal_nota": kw.tanggal_nota,
            "nama_toko": kw.nama_toko,
            "npwp_toko": kw.npwp_toko,
            "alamat_toko": kw.alamat_toko,

            # Data transaksi
            "uraian": uraian_list,

            # ✅ FOTO TERKELOMPOK
            # contoh:
            # {
            #   "barang": [...],
            #   "proses": [...],
            #   "sesudah": [...]
            # }
            "foto_bukti": foto_grouped
        }
    }



@api_router.delete("/kwitansi/{kwitansi_id}/foto/{foto_index}")
async def delete_foto_kwitansi(
    kwitansi_id: int,
    foto_index: int,
    db: AsyncSession = Depends(get_db)
):
    # =========================
    # 1. Ambil kwitansi
    # =========================
    result = await db.execute(
        select(Kwitansi).where(Kwitansi.id == kwitansi_id)
    )
    kw = result.scalar_one_or_none()

    if not kw:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # =========================
    # 2. Parse foto_bukti
    # =========================
    try:
        foto_list = json.loads(kw.foto_bukti or "[]")
    except Exception:
        foto_list = []

    if not isinstance(foto_list, list):
        foto_list = []

    if foto_index < 0 or foto_index >= len(foto_list):
        raise HTTPException(status_code=400, detail="Index foto tidak valid")

    # =========================
    # 3. Ambil path foto
    # =========================
    foto_item = foto_list[foto_index]

    # Support data lama & baru
    if isinstance(foto_item, str):
        foto_path = foto_item
    elif isinstance(foto_item, dict):
        foto_path = foto_item.get("path")
    else:
        foto_path = None

    # =========================
    # 4. Hapus file fisik
    # =========================
    if foto_path:
        # Pastikan path relatif ke root project
        full_path = Path(foto_path)

        if full_path.exists() and full_path.is_file():
            try:
                full_path.unlink()
            except Exception as e:
                logger.error(f"Gagal hapus file {full_path}: {e}")

    # =========================
    # 5. Hapus dari list & update DB
    # =========================
    foto_list.pop(foto_index)
    kw.foto_bukti = json.dumps(foto_list)

    await db.commit()

    # =========================
    # 6. Response
    # =========================
    return {
        "status": "success",
        "message": "Foto berhasil dihapus",
        "sisa_foto": len(foto_list)
    }


@api_router.post("/kwitansi/{kwitansi_id}/update-detail")
async def update_kwitansi_detail(kwitansi_id: int, payload: dict, db: AsyncSession = Depends(get_db)):
    # 1. Ambil data kwitansi
    result = await db.execute(select(Kwitansi).where(Kwitansi.id == kwitansi_id))
    kw = result.scalars().first()
    if not kw:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # 2. Ambil semua transaksi asli untuk backup data keuangan per baris
    trans_res = await db.execute(
        select(Transaksi).where(Transaksi.no_bukti == kw.no_bukti).order_by(Transaksi.id.asc())
    )
    old_transaksi = trans_res.scalars().all()
    
    if not old_transaksi:
        raise HTTPException(status_code=404, detail="Data transaksi tidak ditemukan")

    # 3. Update data identitas di tabel Kwitansi
    kw.nama_kegiatan = payload.get("nama_kegiatan", kw.nama_kegiatan)
    kw.no_bast = payload.get("no_bast", "")

    # --- PARSE TANGGAL DARI FRONTEND TYPE="DATE" ---
    tanggal_nota_str = payload.get("tanggal_nota", "")
    if tanggal_nota_str:
        try:
            # Frontend <input type="date"> selalu mengirim "YYYY-MM-DD"
            kw.tanggal_nota = datetime.strptime(tanggal_nota_str, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal tidak valid")
    # Jika kosong, biarkan tetap di DB (tidak overwrite)
    
    kw.nama_toko = payload.get("nama_toko", "")
    kw.npwp_toko = payload.get("npwp_toko", "")
    kw.alamat_toko = payload.get("alamat_toko", "")

    # 4. Sinkronisasi Uraian tanpa menghilangkan nominal
    new_uraian_list = payload.get("uraian") or []

    # Hapus transaksi lama tapi backup di memori
    await db.execute(delete(Transaksi).where(Transaksi.no_bukti == kw.no_bukti))
    
    for i, item in enumerate(new_uraian_list):
        source = old_transaksi[i] if i < len(old_transaksi) else None
        
        # Gunakan tanggal kwitansi yang sudah diparse, fallback aman ke kw.tanggal_nota
        t = Transaksi(
            no_bukti=kw.no_bukti,
            kode_kegiatan=kw.kode_kegiatan,
            tanggal=kw.tanggal_nota,
            uraian=item.get("nama_barang") or "",
            volume=item.get("volume") or 0,
            satuan=item.get("satuan") or "",
            harga_satuan=item.get("harga_satuan") or 0,
            kode_rekening=source.kode_rekening if source else old_transaksi[0].kode_rekening,
            penerimaan=source.penerimaan if source else 0,
            pengeluaran=source.pengeluaran if source else 0,
            saldo=source.saldo if source else 0
        )
        db.add(t)

    db.add(kw)
    await db.commit()
    
    return {
        "status": "success",
        "message": f"Kwitansi '{kw.nama_kegiatan}' berhasil diperbarui tanpa mengubah nilai pengeluaran."
    }


@api_router.post("/kwitansi/{kwitansi_id}/upload-foto")
async def upload_foto_kwitansi(
    kwitansi_id: int,
    files: List[UploadFile] = File(...),
    jenis_foto: str = Form("barang"),  # ⬅⬅⬅ TAMBAHAN PENTING
    db: AsyncSession = Depends(get_db)
):
    # =========================
    # 1. Validasi jenis foto
    # =========================
    JENIS_VALID = ["barang", "kegiatan", "sebelum", "proses", "sesudah"]
    if jenis_foto not in JENIS_VALID:
        raise HTTPException(
            status_code=400,
            detail=f"Jenis foto tidak valid. Pilihan: {', '.join(JENIS_VALID)}"
        )

    # =========================
    # 2. Validasi file
    # =========================
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file diupload")

    if len(files) > 15:
        raise HTTPException(status_code=400, detail="Maksimal 15 foto")

    # =========================
    # 3. Ambil kwitansi
    # =========================
    result = await db.execute(
        select(Kwitansi).where(Kwitansi.id == kwitansi_id)
    )
    kwitansi = result.scalar_one_or_none()

    if not kwitansi:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # =========================
    # 4. Folder upload
    # =========================
    kw_dir = UPLOAD_KWITANSI_DIR / str(kwitansi_id)
    kw_dir.mkdir(parents=True, exist_ok=True)

    foto_baru = []

    # =========================
    # 5. Simpan file
    # =========================
    for file in files:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in [".jpg", ".jpeg", ".png"]:
            continue

        filename = f"{uuid.uuid4().hex}{ext}"
        filepath = kw_dir / filename

        with open(filepath, "wb") as f:
            f.write(await file.read())

        foto_baru.append({
            "path": f"uploads/kwitansi/{kwitansi_id}/{filename}",
            "jenis": jenis_foto
        })

    if not foto_baru:
        raise HTTPException(status_code=400, detail="Tidak ada file gambar valid")

    # =========================
    # 6. Gabung dengan foto lama (AMAN)
    # =========================
    foto_lama = []
    try:
        raw = json.loads(kwitansi.foto_bukti or "[]")
        for item in raw:
            if isinstance(item, str):
                foto_lama.append({
                    "path": item,
                    "jenis": "barang"  # default aman
                })
            elif isinstance(item, dict):
                foto_lama.append({
                    "path": item.get("path"),
                    "jenis": item.get("jenis", "barang")
                })
    except Exception:
        pass

    merged = foto_lama + foto_baru
    kwitansi.foto_bukti = json.dumps(merged)
    await db.commit()

    return {"status": "success", "message": "Foto berhasil diupload", "uploaded": len(foto_baru)}

# --- 1. SCHEMAS (Pydantic) ---
class AIImageRequest(BaseModel):
    prompt_items: str  # Contoh: "Laptop Asus ROG" atau "Rapat koordinasi guru"
    jenis_foto: str    # Valid: "barang", "kegiatan", "sebelum", "proses", "sesudah"

# --- 2. HELPER: PROMPT ENGINEER ---
def build_professional_prompt(item: str, jenis: str) -> str:
    """
    Membuat prompt spesifik agar hasil terlihat seperti foto dokumentasi asli Indonesia.
    Menggunakan model FLUX untuk realisme.
    """
    base_style = (
        "realistic photo taken with smartphone camera, "
        "poor indoor lighting, authentic amateur photography, "
        "no filter, slight motion blur, documentary style, "
        "high texture, taken in Indonesia, "
    )
    
    if jenis == "kegiatan" or jenis == "proses":
        # Prompt untuk aktivitas (ada orang, suasana kantor)
        return (
            f"{base_style} "
            f"wide shot of people doing {item}, "
            f"people wearing Indonesian civil servant uniform or Batik shirt, "
            f"busy office background, piles of paper on tables, "
            f"candid shot, messy real environment, "
            f"faces slightly blurred for privacy."
        )
    else:
        # Prompt untuk barang (benda mati di meja)
        return (
            f"{base_style} "
            f"close up photo of {item}, "
            f"placed on a scratched wooden office desk, "
            f"background is a blurry administration room wall, "
            f"top down angle or eye level, "
            f"sharp focus on the object, real object proportions."
        )

# --- 3. ENDPOINT ---
@api_router.post("/kwitansi/{kwitansi_id}/generate-ai-image")
async def generate_ai_image(
    kwitansi_id: int,
    request: AIImageRequest,
    db: AsyncSession = Depends(get_db)  # Pastikan ini sesuai dependency injection Anda
):
    # A. Validasi Input
    JENIS_VALID = ["barang", "kegiatan", "sebelum", "proses", "sesudah"]
    if request.jenis_foto not in JENIS_VALID:
        raise HTTPException(status_code=400, detail=f"Jenis foto harus salah satu dari: {JENIS_VALID}")

    # B. Cek Data di Database
    result = await db.execute(select(Kwitansi).where(Kwitansi.id == kwitansi_id))
    kwitansi = result.scalar_one_or_none()
    
    if not kwitansi:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # C. Siapkan Folder
    kw_dir = UPLOAD_KWITANSI_DIR / str(kwitansi_id)
    kw_dir.mkdir(parents=True, exist_ok=True)

    try:
        # D. Generate Prompt & URL
        # Kita gunakan model 'flux' yang jauh lebih realistis daripada default
        clean_prompt = build_professional_prompt(request.prompt_items, request.jenis_foto)
        encoded_prompt = quote(clean_prompt.strip())
        
        # Seed acak agar gambar selalu unik
        seed = uuid.uuid4().int % 100000
        
        # URL Pollinations dengan Model Flux
        ai_url = (
            f"https://image.pollinations.ai/prompt/{encoded_prompt}"
            f"?width=1024&height=1024"
            f"&model=flux"      # <--- KUNCI REALISME
            f"&seed={seed}"
            f"&nologo=true"     # Hilangkan watermark
            f"&enhance=false"   # False agar tidak terlalu 'cantik' (biar natural)
        )

        # E. Request ke AI Server (Non-blocking I/O)
        # Timeout diset 60s karena model Flux kadang butuh waktu render
        async with httpx.AsyncClient() as client:
            response = await client.get(ai_url, timeout=60.0)

        # Fallback jika model Flux gagal/sibuk (biasanya return text error bukan image)
        if response.status_code != 200 or "image" not in response.headers.get("content-type", ""):
            # Coba sekali lagi tanpa parameter model (pakai default)
            print("Flux model busy, retrying with default...")
            fallback_url = ai_url.replace("&model=flux", "")
            async with httpx.AsyncClient() as client:
                response = await client.get(fallback_url, timeout=60.0)
                
            if response.status_code != 200:
                 raise HTTPException(status_code=502, detail="AI Service sedang sibuk, coba lagi nanti.")

        # F. Simpan Gambar ke Disk
        filename = f"ai_{request.jenis_foto}_{uuid.uuid4().hex[:8]}.jpg"
        filepath = kw_dir / filename

        with open(filepath, "wb") as f:
            f.write(response.content)

        # G. Update Database (JSON Append)
        # Logika merge foto lama dan baru
        foto_list = []
        if kwitansi.foto_bukti:
            try:
                raw_data = json.loads(kwitansi.foto_bukti)
                # Normalisasi data lama ke format dict
                for item in raw_data:
                    if isinstance(item, str):
                        foto_list.append({"path": item, "jenis": "barang", "source": "manual"})
                    elif isinstance(item, dict):
                        foto_list.append(item)
            except json.JSONDecodeError:
                foto_list = []

        # Tambah foto baru
        new_photo_entry = {
            "path": f"uploads/kwitansi/{kwitansi_id}/{filename}",
            "jenis": request.jenis_foto,
            "source": "ai",
            "prompt_used": request.prompt_items # Opsional: simpan prompt buat history
        }
        foto_list.append(new_photo_entry)

        kwitansi.foto_bukti = json.dumps(foto_list)
        await db.commit()
        await db.refresh(kwitansi)

        return {
            "status": "success",
            "message": "Gambar berhasil digenerate",
            "data": new_photo_entry
        }

    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Waktu habis saat menghubungi AI server")
    except Exception as e:
        await db.rollback()
        # Print error log di terminal server
        print(f"ERROR GENERATE AI: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}")

@api_router.post("/kwitansi/manual")
async def create_manual_kwitansi(
    no_bukti: str = Form(...),
    kode_kegiatan: str = Form(""),
    kode_rekening: str = Form(""),
    tanggal: str = Form(None),
    tanggal_nota: str = Form(""),
    no_bast: str = Form(""),
    nama_toko: str = Form(""),
    npwp_toko: str = Form(""),
    alamat_toko: str = Form(""),
    uraian_json: str = Form(...),
    files: List[UploadFile] = File(None),
    jenis_foto: str = Form("barang"),
    db: AsyncSession = Depends(get_db)
):
    """Create a kwitansi + transaksi from manual input. Accepts multipart form: uraian_json (JSON array) and optional files (foto).
    uraian_json example: [{"nama_barang":"...","volume":1,"satuan":"...","harga_satuan":1000,"pengeluaran":1000}, ...]
    """
    try:
        # Basic validation
        if not no_bukti or not no_bukti.strip():
            raise HTTPException(status_code=400, detail="no_bukti is required")

        try:
            uraian = json.loads(uraian_json)
            if not isinstance(uraian, list) or len(uraian) == 0:
                raise ValueError()
        except Exception:
            raise HTTPException(status_code=400, detail="uraian_json harus berupa JSON array yang valid dan tidak kosong")

        # Ensure kode_rekening maps to a master (optional)
        master_rek = None
        if kode_rekening:
            res = await db.execute(select(MasterRekeningBelanja).where(MasterRekeningBelanja.kode_rekening_belanja == kode_rekening))
            master_rek = res.scalar_one_or_none()

        # Insert transaksi rows
        total_pengeluaran = 0.0
        transaksi_objs = []
        for item in uraian:
            nama_barang = item.get("nama_barang") or item.get("uraian") or ""
            volume = float(item.get("volume") or 0)
            satuan = item.get("satuan") or ""
            harga_satuan = float(item.get("harga_satuan") or 0)
            pengeluaran = float(item.get("pengeluaran") or 0)

            t = Transaksi(
                tanggal=tanggal or (item.get("tanggal") or ""),
                kode_kegiatan=kode_kegiatan or "",
                kode_rekening=kode_rekening or "",
                no_bukti=no_bukti,
                uraian=nama_barang,
                volume=volume,
                satuan=satuan,
                harga_satuan=harga_satuan,
                penerimaan=0,
                pengeluaran=pengeluaran,
            )
            db.add(t)
            transaksi_objs.append(t)
            total_pengeluaran += pengeluaran

        # Commit transactions to get their insertion timestamps
        await db.commit()

        # Determine tanggal/thp/tahun for kwitansi
        tanggal_for_kw = None
        for t in transaksi_objs:
            if t.tanggal:
                try:
                    parsed = parse_date(t.tanggal, dayfirst=True)
                    tanggal_for_kw = parsed if not tanggal_for_kw or parsed < tanggal_for_kw else tanggal_for_kw
                except Exception:
                    pass
        if not tanggal_for_kw and tanggal:
            try:
                tanggal_for_kw = parse_date(tanggal, dayfirst=True)
            except Exception:
                tanggal_for_kw = None

        if tanggal_for_kw:
            bulan = tanggal_for_kw.month
            tahun = tanggal_for_kw.year
            thp = "THP I" if bulan <= 6 else "THP II"
        else:
            thp = ""
            tahun = ""

        # Resolve nama_kegiatan from master_kegiatan if possible
        nama_kegiatan = ""
        if kode_kegiatan:
            r = await db.execute(select(MasterKegiatan).where(MasterKegiatan.kode_kegiatan == kode_kegiatan))
            mk = r.scalar_one_or_none()
            if mk:
                nama_kegiatan = mk.nama_kegiatan

        nomor_kwitansi = f"{no_bukti}/{kode_kegiatan or '-'} /{thp}/{tahun}" if thp and tahun else None

        kw = Kwitansi(
            nomor_kwitansi=nomor_kwitansi or (no_bukti),
            no_bukti=no_bukti,
            kode_kegiatan=kode_kegiatan or "",
            nama_kegiatan=nama_kegiatan or "",
            tanggal=(tanggal_for_kw.strftime("%d-%m-%Y") if tanggal_for_kw else (tanggal or "")),
            thp=thp,
            tahun=str(tahun) if tahun else "",
            jumlah=total_pengeluaran,
            tanggal_nota=tanggal_nota or "",
            no_bast=no_bast or "",
            nama_toko=nama_toko or "",
            npwp_toko=npwp_toko or "",
            alamat_toko=alamat_toko or "",
            foto_bukti=json.dumps([]),
        )

        db.add(kw)
        await db.commit()
        await db.refresh(kw)

        # Handle file uploads (if any)
        foto_baru = []
        if files:
            # reuse upload logic
            kw_dir = UPLOAD_KWITANSI_DIR / str(kw.id)
            kw_dir.mkdir(parents=True, exist_ok=True)
            for file in files:
                ext = os.path.splitext(file.filename)[1].lower()
                if ext not in [".jpg", ".jpeg", ".png"]:
                    continue
                filename = f"{uuid.uuid4().hex}{ext}"
                filepath = kw_dir / filename
                with open(filepath, "wb") as f:
                    f.write(await file.read())
                foto_baru.append({"path": f"uploads/kwitansi/{kw.id}/{filename}", "jenis": jenis_foto})

            if foto_baru:
                raw_old = []
                try:
                    raw_old = json.loads(kw.foto_bukti or "[]")
                except Exception:
                    raw_old = []

                merged = []
                for it in raw_old:
                    if isinstance(it, str):
                        merged.append({"path": it, "jenis": "barang"})
                    elif isinstance(it, dict):
                        merged.append({"path": it.get("path"), "jenis": it.get("jenis", "barang")})
                merged.extend(foto_baru)
                kw.foto_bukti = json.dumps(merged)
                await db.commit()

        return {"status": "success", "message": "Kwitansi & transaksi berhasil dibuat", "kwitansi_id": kw.id, "nomor_kwitansi": kw.nomor_kwitansi}

    except Exception as e:
        await db.rollback()
        logger.error(f"Error creating manual kwitansi: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating manual kwitansi: {str(e)}")
    # =========================
    # 1. Validasi jenis foto
    # =========================
    JENIS_VALID = ["barang", "kegiatan", "sebelum", "proses", "sesudah"]
    if jenis_foto not in JENIS_VALID:
        raise HTTPException(
            status_code=400,
            detail=f"Jenis foto tidak valid. Pilihan: {', '.join(JENIS_VALID)}"
        )

    # =========================
    # 2. Validasi file
    # =========================
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file diupload")

    if len(files) > 15:
        raise HTTPException(status_code=400, detail="Maksimal 15 foto")

    # =========================
    # 3. Ambil kwitansi
    # =========================
    result = await db.execute(
        select(Kwitansi).where(Kwitansi.id == kwitansi_id)
    )
    kwitansi = result.scalar_one_or_none()

    if not kwitansi:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # =========================
    # 4. Folder upload
    # =========================
    kw_dir = UPLOAD_KWITANSI_DIR / str(kwitansi_id)
    kw_dir.mkdir(parents=True, exist_ok=True)

    foto_baru = []

    # =========================
    # 5. Simpan file
    # =========================
    for file in files:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in [".jpg", ".jpeg", ".png"]:
            continue

        filename = f"{uuid.uuid4().hex}{ext}"
        filepath = kw_dir / filename

        with open(filepath, "wb") as f:
            f.write(await file.read())

        foto_baru.append({
            "path": f"uploads/kwitansi/{kwitansi_id}/{filename}",
            "jenis": jenis_foto
        })

    if not foto_baru:
        raise HTTPException(status_code=400, detail="Tidak ada file gambar valid")

    # =========================
    # 6. Gabung dengan foto lama (AMAN)
    # =========================
    foto_lama = []
    try:
        raw = json.loads(kwitansi.foto_bukti or "[]")
        for item in raw:
            if isinstance(item, str):
                foto_lama.append({
                    "path": item,
                    "jenis": "barang"  # default aman
                })
            elif isinstance(item, dict):
                foto_lama.append({
                    "path": item.get("path"),
                    "jenis": item.get("jenis", "barang")
                })
    except Exception:
        pass

    kwitansi.foto_bukti = json.dumps(foto_lama + foto_baru)
    await db.commit()

    # =========================
    # 7. Response
    # =========================
    return {
        "status": "success",
        "jenis": jenis_foto,
        "jumlah": len(foto_baru),
        "files": foto_baru
    }


# --- ENDPOINT BARU UNTUK LOGO & PEMDA ---

@api_router.post("/save-pemda")
async def save_pemda(
    nama_pemda: str = Form(...),
    nama_sekolah: str = Form(...),
    alamat_sekolah: str = Form(...),
    tempat_surat: str = Form(""),
    logo_pemda: UploadFile = File(None),
    logo_sekolah: UploadFile = File(None),
    db: AsyncSession = Depends(get_db)
):
    """Save Pemda data and logos using multipart/form-data"""
    try:
        # 1. Ambil data pengaturan ID 1
        stmt = select(Pengaturan).limit(1)
        result = await db.execute(stmt)
        pengaturan = result.scalar_one_or_none()

        if not pengaturan:
            pengaturan = Pengaturan()
            db.add(pengaturan)

        # 2. Update Field Teks
        pengaturan.nama_pemda = nama_pemda
        pengaturan.nama_sekolah = nama_sekolah
        pengaturan.alamat_sekolah = alamat_sekolah
        pengaturan.tempat_surat = tempat_surat

        # Log DB location & values for diagnostics
        try:
            from database import DB_PATH
            logger.info(f"save_pemda: menulis ke DB_PATH={DB_PATH} (exists={DB_PATH.exists()}) | tempat_surat='{pengaturan.tempat_surat}' | pengaturan.id={getattr(pengaturan, 'id', None)}")
        except Exception:
            logger.info("save_pemda: tidak dapat membaca DB_PATH untuk logging")

        # 3. Handle Upload Logo Pemda (simpan ke folder uploads/logos)
        if logo_pemda and logo_pemda.filename:
            ext = os.path.splitext(logo_pemda.filename)[1]
            filename = f"logo_pemda{ext}"
            file_path = UPLOAD_LOGO_DIR / filename
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(logo_pemda.file, buffer)
            pengaturan.logo_pemda = filename # Simpan nama filenya saja
            logger.info("save_pemda: logo_pemda disimpan di %s", file_path)

        # 4. Handle Upload Logo Sekolah (simpan ke folder uploads/logos)
        if logo_sekolah and logo_sekolah.filename:
            ext = os.path.splitext(logo_sekolah.filename)[1]
            filename = f"logo_sekolah{ext}"
            file_path = UPLOAD_LOGO_DIR / filename
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(logo_sekolah.file, buffer)
            pengaturan.logo_sekolah = filename
            logger.info("save_pemda: logo_sekolah disimpan di %s", file_path)

        await db.commit()
        return {"status": "success", "message": "Data berhasil diperbarui"}

    except Exception as e:
        await db.rollback()
        logger.error(f"Error saving pemda: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/get-pemda")
async def get_pemda(
    request: Request,
    db: AsyncSession = Depends(get_db)
):
    try:
        stmt = select(Pengaturan).limit(1)
        result = await db.execute(stmt)
        p = result.scalar_one_or_none()

        if not p:
            return {"status": "success", "data": None}

        base_url = str(request.base_url).rstrip("/")

        return {
            "status": "success",
            "data": {
                "nama_pemda": p.nama_pemda,
                "nama_sekolah": p.nama_sekolah,
                "alamat_sekolah": p.alamat_sekolah,
                "tempat_surat": p.tempat_surat or "",
                "logo_pemda": (
                    f"{base_url}/uploads/logos/{p.logo_pemda}"
                    if p.logo_pemda else None
                ),
                "logo_sekolah": (
                    f"{base_url}/uploads/logos/{p.logo_sekolah}"
                    if p.logo_sekolah else None
                )
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/db-info")
async def db_info(request: Request):
    """Diagnostic endpoint: returns DB path, existence and pengaturan row"""
    try:
        from database import DB_PATH
        conn = None
        import sqlite3
        from pathlib import Path
        p = Path(DB_PATH)
        exists = p.exists()
        rows = None
        peng = None
        if exists:
            conn = sqlite3.connect(str(p))
            cur = conn.cursor()
            cur.execute("PRAGMA table_info('pengaturan')")
            cols = [r[1] for r in cur.fetchall()]
            cur.execute("SELECT * FROM pengaturan LIMIT 1")
            row = cur.fetchone()
            if row:
                peng = dict(zip(cols, row))
        if conn:
            conn.close()
        return {
            "status": "success",
            "db_path": str(p),
            "db_exists": exists,
            "pengaturan": peng
        }
    except Exception as e:
        logger.exception("db_info gagal: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

def format_rupiah(value):
    """Format float ke string Rupiah"""
    try:
        return f"Rp {value:,.0f}".replace(",", ".")
    except:
        return "Rp 0"

def format_tanggal_indo(tanggal_str):
    """Mengubah 'YYYY-MM-DD' menjadi object datetime dan string indo"""
    # Asumsi format tanggal di database string 'YYYY-MM-DD'
    try:
        dt = datetime.strptime(tanggal_str, '%Y-%m-%d')
    except ValueError:
        # Fallback jika formatnya lain atau kosong
        dt = datetime.now()
    
    bulan_indo = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", 
                  "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    
    hari_indo = {
        'Monday': 'Senin', 'Tuesday': 'Selasa', 'Wednesday': 'Rabu',
        'Thursday': 'Kamis', 'Friday': 'Jumat', 'Saturday': 'Sabtu', 'Sunday': 'Minggu'
    }
    
    return {
        'obj': dt,
        'hari': hari_indo[dt.strftime('%A')],
        'tanggal': str(dt.day),
        'bulan': bulan_indo[dt.month],
        'tahun': str(dt.year),
        'full': f"{dt.day} {bulan_indo[dt.month]} {dt.year}",
        'terbilang_tanggal': num2words(dt.day, lang='id').title(),
        'terbilang_tahun': num2words(dt.year, lang='id').title()
    }

def generate_bast_pdf(kwitansi, pengaturan, transaksi_list):
    # --- 4. OLAH TANGGAL ---
    raw_date_str = kwitansi.tanggal_nota if kwitansi.tanggal_nota else kwitansi.tanggal

    try:
        if not raw_date_str:
            tgl_obj = datetime.now()
        elif "/" in str(raw_date_str):
            tgl_obj = datetime.strptime(raw_date_str, "%d/%m/%Y")
        elif "-" in str(raw_date_str) and str(raw_date_str).find("-") == 4:
            tgl_obj = datetime.strptime(raw_date_str, "%Y-%m-%d")
        elif "-" in str(raw_date_str):
            tgl_obj = datetime.strptime(raw_date_str, "%d-%m-%Y")
        else:
            tgl_obj = datetime.now()
    except Exception:
        tgl_obj = datetime.now()

    # Helper untuk tanggal Indonesia
    tgl_bast = format_tanggal_indo(tgl_obj.strftime("%Y-%m-%d"))

    # 5. Data Header untuk PDF
    data = {
        "pemda": (pengaturan.nama_pemda or "").upper(),
        "sekolah": (pengaturan.nama_sekolah or "").upper(),
        "alamat": pengaturan.alamat_sekolah or "",
        "nama_kepala_sekolah": pengaturan.nama_kepala_sekolah or "",
        "nip_kepala_sekolah": pengaturan.nip_kepala_sekolah or "",
        "nama_bendahara": pengaturan.nama_bendahara or "",
        "nip_bendahara": pengaturan.nip_bendahara or "",
        "logo_pemda": resolve_logo(pengaturan.logo_pemda),
        "logo_sekolah": resolve_logo(pengaturan.logo_sekolah),
        "nomor_surat": kwitansi.no_bast or "-",
        "hari_ini": tgl_bast.get("hari", ""),
        "tgl_terbilang": tgl_bast.get("terbilang_tanggal", ""),
        "bulan_indo": tgl_bast.get("bulan", ""),
        "tahun_terbilang": tgl_bast.get("terbilang_tahun", ""),
        "tgl_full": tgl_bast.get("full", ""), 
        "nama_penerima": pengaturan.nama_pengurus_barang or "",
        "nip_penerima": pengaturan.nip_pengurus_barang or "",
        "jabatan_penerima": "Pengurus Barang",
        "nama_penyedia": kwitansi.nama_toko or "",
        "tgl_nota_full": tgl_bast.get("full", ""),
        "kota": pengaturan.tempat_surat or "Jember",
    }

    # 6. Item barang & Hitung Total
    items = []
    total_volume = 0.0
    grand_total_num = 0.0

    def _parse_number(value):
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if s == "":
            return 0.0
        try:
            return float(s)
        except Exception:
            pass
        try:
            if s.count('.') > 0 and s.count(',') == 0:
                return float(s.replace(',', ''))
            if s.count(',') > 0 and s.count('.') == 0:
                return float(s.replace(',', '.'))
            if s.count('.') > 0 and s.count(',') > 0:
                return float(s.replace('.', '').replace(',', '.'))
        except Exception:
            pass
        s2 = re.sub(r'[^0-9.-]', '', s)
        try:
            return float(s2)
        except Exception:
            logger.warning("Gagal parse angka: %s", s)
            return 0.0

    for idx, trx in enumerate(transaksi_list, start=1):
        try:
            vol = _parse_number(trx.volume)
            hrg = _parse_number(trx.harga_satuan)
        except Exception as e:
            logger.exception("Error parsing angka", e)
            vol = 0.0
            hrg = 0.0

        sub = vol * hrg
        items.append({
            "no": idx,
            "uraian": trx.uraian or "-",
            "harga_satuan": format_rupiah(hrg),
            "satuan": trx.satuan or "-",
            "volume": int(vol) if isinstance(vol, float) and vol.is_integer() else vol,
            "jumlah": format_rupiah(sub)
        })
        total_volume += vol
        grand_total_num += sub

    total_volume_fmt = int(total_volume) if isinstance(total_volume, float) and total_volume.is_integer() else total_volume
    grand_total_str = format_rupiah(grand_total_num)
    
    # 7. Generate PDF
    pdf = FPDF(orientation='P', unit='mm', format=(215, 330))
    pdf.set_left_margin(15)
    pdf.set_right_margin(15)
    pdf.set_top_margin(15)

    # --- HALAMAN TANDA TERIMA PEMBAYARAN ---
    pdf.add_page()
    
    # --- KOP SURAT (Page 1) ---
    tinggi_logo = 18
    
    # >>> LOGO PEMDA (Page 1) <<<
    logo_pemda_path = data.get('logo_pemda') or ''
    if logo_pemda_path:
        try:
            p_logo = Path(logo_pemda_path)
            if p_logo.exists():
                pdf.image(str(p_logo), x=15, y=22, h=tinggi_logo)
        except Exception:
            pass
            
    # Logo Sekolah
    logo_sekolah_path = data.get('logo_sekolah') or ''
    if logo_sekolah_path:
        try:
            p_logo_s = Path(logo_sekolah_path)
            if p_logo_s.exists():
                pdf.image(str(p_logo_s), x=182, y=22, h=tinggi_logo)
        except Exception:
            pass

    pdf.set_y(22) 
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 6, data['pemda'], align="C", ln=1)
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 7, data['sekolah'], align="C", ln=1)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(0, 5, data['alamat'], align="C", ln=1)

    pdf.set_line_width(0.6)
    pdf.line(15, 45, 200, 45)
    pdf.ln(15)

    # --- JUDUL HALAMAN ---
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 6, "TANDA TERIMA PEMBAYARAN", align="C", ln=1)
    nama_kegiatan_upper = (kwitansi.nama_kegiatan or "").upper()
    pdf.multi_cell(0, 6, nama_kegiatan_upper, align="C")
    pdf.cell(0, 6, f"Nomor Kwitansi : {kwitansi.nomor_kwitansi or '-'}", align="C", ln=1)
    pdf.ln(5)

    # --- TABEL BARANG / JASA ---
    w = [10, 65, 30, 20, 30, 30]
    pdf.set_line_width(0.15)
    pdf.set_font("helvetica", "B", 10)
    headers = ["No", "Rincian Barang/Jasa", "Volume", "Satuan", "Harga", "Jumlah"]
    
    for i, h in enumerate(headers):
        pdf.cell(w[i], 8, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("helvetica", "", 10)
    for item in items:
        start_y = pdf.get_y()
        x_start = pdf.get_x()
        pdf.set_xy(x_start + w[0], start_y)
        pdf.multi_cell(w[1], 8, item['uraian'], border=1, align="L")
        end_y = pdf.get_y()
        h_row = end_y - start_y
        pdf.set_xy(x_start, start_y)
        pdf.cell(w[0], h_row, str(item['no']), border=1, align="C")
        pdf.set_xy(x_start + w[0] + w[1], start_y)
        pdf.cell(w[2], h_row, str(item['volume']), border=1, align="C")
        pdf.cell(w[3], h_row, item['satuan'], border=1, align="C")
        pdf.cell(w[4], h_row, item['harga_satuan'], border=1, align="R")
        pdf.cell(w[5], h_row, item['jumlah'], border=1, align="R")
        pdf.set_y(end_y)

    pdf.set_fill_color(191, 191, 191)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(sum(w[0:4]), 8, "Jumlah Total", border=1, align="R", fill=True)
    pdf.cell(w[4], 8, "", border=1, fill=True)
    pdf.cell(w[5], 8, grand_total_str, border=1, align="R", fill=True)
    pdf.ln(15)

    # --- TANDA TANGAN ---
    pdf.ln(10)
    pdf.set_font("helvetica", "", 11)
    pdf.set_x(135)
    pdf.cell(60, 5, f"{data['kota']}, {data['tgl_full']}", align="C", ln=1)
    pdf.set_x(15)
    pdf.cell(60, 5, "Bendahara BOSP", align="C")
    pdf.set_x(135)
    pdf.cell(60, 5, "Penerima", align="C")
    pdf.ln(20)
    pdf.set_font("helvetica", "BU", 11)
    pdf.set_x(15)
    pdf.cell(60, 5, data['nama_bendahara'], align="C")
    pdf.set_x(135)
    pdf.cell(60, 5, data['nama_penyedia'], align="C")
    pdf.ln(5)
    pdf.set_font("helvetica", "", 11)
    pdf.set_x(15)
    pdf.cell(60, 5, f"NIP. {data['nip_bendahara']}", align="C")
    pdf.ln(15)
    pdf.set_x(75) 
    pdf.cell(60, 5, "Mengetahui,", align="C", ln=1)
    pdf.set_x(75)
    pdf.cell(60, 5, "Kepala Sekolah", align="C")
    pdf.ln(20)
    pdf.set_font("helvetica", "BU", 11)
    pdf.set_x(75)
    pdf.cell(60, 5, data['nama_kepala_sekolah'], align="C")
    pdf.ln(5)
    pdf.set_font("helvetica", "", 11)
    pdf.set_x(75)
    pdf.cell(60, 5, f"NIP. {data['nip_kepala_sekolah']}", align="C")

    # --- HALAMAN BAST (PAGES 2) ---
    pdf.add_page()
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 5, "FORMAT 2", align="R", ln=1)
    pdf.ln(2)
    
    # >>> LOGO PEMDA (Page 2) <<<
    logo_pemda_path = data.get('logo_pemda') or ''
    if logo_pemda_path:
        try:
            p_logo = Path(logo_pemda_path)
            if p_logo.exists():
                pdf.image(str(p_logo), x=15, y=22, h=tinggi_logo)
        except Exception:
            pass
            
    # Logo Sekolah (Page 2)
    logo_sekolah_path = data.get('logo_sekolah') or ''
    if logo_sekolah_path:
        try:
            p_logo_s = Path(logo_sekolah_path)
            if p_logo_s.exists():
                pdf.image(str(p_logo_s), x=182, y=22, h=tinggi_logo)
        except Exception:
            pass

    pdf.set_y(22)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 6, data['pemda'], align="C", ln=1)
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 7, data['sekolah'], align="C", ln=1)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(0, 5, data['alamat'], align="C", ln=1)
    pdf.set_line_width(0.6)
    pdf.line(15, 45, 200, 45) 
    pdf.ln(15)
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 5, "BERITA ACARA SERAH TERIMA BARANG", align="C", ln=1)
    pdf.cell(0, 5, "PENGADAAN / PEROLEHAN", align="C", ln=1)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(0, 5, f"nomor: {data['nomor_surat']}", align="C", ln=1)
    pdf.ln(8)
    pdf.set_font("helvetica", "", 11)
    teks_pembuka = f"Pada hari ini {data['hari_ini']} tanggal {data['tgl_terbilang']} bulan {data['bulan_indo']} tahun {data['tahun_terbilang']} yang bertanda tangan di bawah ini :"
    pdf.multi_cell(0, 5, teks_pembuka, align="J")
    pdf.ln(3)
    
    # Detail BAST
    pdf.cell(30, 6, "Nama")
    pdf.cell(0, 6, f": {data['nama_penerima']}", ln=1)
    pdf.cell(30, 6, "Jabatan")
    pdf.cell(0, 6, f": {data['jabatan_penerima']}", ln=1)
    pdf.ln(3)
    teks_isi = f"Telah menerima barang persediaan yang diserahkan oleh {data['nama_penyedia']} sesuai dengan Faktur Pembelian / Nota Pembelian tanggal {data['tgl_nota_full']} sebagaimana daftar terlampir."
    pdf.multi_cell(0, 5, teks_isi, align="J")
    pdf.ln(3)
    pdf.cell(0, 5, "Daftar barang yang diterima sebagai berikut :", ln=1)
    pdf.ln(2)

    # TABEL BAST
    w_bast = [10, 65, 30, 20, 15, 30, 15]
    pdf.set_line_width(0.15)
    pdf.set_font("helvetica", "B", 10)
    headers_bast = ["No", "Uraian Nama Barang", "Harga Satuan", "Satuan", "Volume", "Jumlah", "Ket"]
    for i in range(len(headers_bast)):
        pdf.cell(w_bast[i], 8, headers_bast[i], border=1, align="C")
    pdf.ln()
    pdf.set_font("helvetica", "I", 8)
    sub_h = ["(a)", "(b)", "(c)", "(d)", "(e)", "(f = c*e)", "(g)"]
    for i in range(len(sub_h)):
        pdf.cell(w_bast[i], 5, sub_h[i], border=1, align="C")
    pdf.ln()
    pdf.set_font("helvetica", "", 10)
    for item in items:
        start_y = pdf.get_y()
        x_start = pdf.get_x()
        pdf.set_xy(x_start + w_bast[0], start_y)
        pdf.multi_cell(w_bast[1], 8, item['uraian'], border=1, align="L")
        end_y = pdf.get_y()
        h_row = end_y - start_y
        pdf.set_xy(x_start, start_y)
        pdf.cell(w_bast[0], h_row, str(item['no']), border=1, align="C")
        pdf.set_xy(x_start + w_bast[0] + w_bast[1], start_y)
        pdf.cell(w_bast[2], h_row, item['harga_satuan'], border=1, align="R")
        pdf.cell(w_bast[3], h_row, item['satuan'], border=1, align="C")
        pdf.cell(w_bast[4], h_row, str(item['volume']), border=1, align="C")
        pdf.cell(w_bast[5], h_row, item['jumlah'], border=1, align="R")
        pdf.cell(w_bast[6], h_row, "", border=1, align="C")
        pdf.set_y(end_y)

    pdf.set_fill_color(191, 191, 191)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(sum(w_bast[0:4]), 8, "Jumlah", border=1, align="R", fill=True)
    pdf.cell(w_bast[4], 8, str(total_volume_fmt), border=1, align="C", fill=True)
    pdf.cell(w_bast[5], 8, grand_total_str, border=1, align="R", fill=True)
    pdf.cell(w_bast[6], 8, "", border=1, fill=True, ln=1)
    pdf.set_fill_color(217, 217, 217)
    pdf.cell(sum(w_bast[0:4]), 8, "Jumlah Total", border=1, align="R", fill=True)
    pdf.cell(w_bast[4], 8, "", border=1, fill=True) 
    pdf.cell(w_bast[5], 8, grand_total_str, border=1, align="R", fill=True)
    pdf.cell(w_bast[6], 8, "", border=1, fill=True, ln=1)

    pdf.set_font("helvetica", "", 11)
    pdf.ln(5)
    pdf.multi_cell(0, 5, "Demikian Berita Acara Serah Terima Barang ini dibuat dalam rangkap 2 (dua) untuk digunakan sebagaimana mestinya.")
    pdf.ln(10)
    pdf.cell(92, 5, "", ln=0)
    pdf.cell(92, 5, f"{data['kota']}, {data['tgl_full']}", align="C", ln=1)
    
    pdf.cell(92, 5, "Yang Menyerahkan", align="C", ln=0)
    pdf.cell(92, 5, "Yang Menerima", align="C", ln=1)
    pdf.cell(92, 5, "Penyedia Barang", align="C", ln=0)
    pdf.cell(92, 5, data['jabatan_penerima'], align="C", ln=1)
    pdf.ln(20) 
    pdf.cell(92, 5, data['nama_penyedia'], align="C", ln=0)
    pdf.set_font("helvetica", "BU", 11)
    pdf.cell(92, 5, data['nama_penerima'], align="C", ln=1)
    pdf.set_font("helvetica", "", 11)
    pdf.cell(92, 5, "", ln=0)
    pdf.cell(92, 5, f"NIP. {data['nip_penerima']}", align="C", ln=1)

    # --- HALAMAN LAMPIRAN FOTO ---
    foto_list = []
    if kwitansi.foto_bukti:
        try:
            foto_list = kwitansi.foto_bukti if isinstance(kwitansi.foto_bukti, list) else json.loads(kwitansi.foto_bukti)
        except Exception:
            foto_list = []

    foto_per_jenis = {k: [] for k in ["barang", "kegiatan", "sebelum", "proses", "sesudah"]}
    for foto in foto_list:
        if isinstance(foto, dict):
            jenis = foto.get("jenis")
            if jenis in foto_per_jenis:
                foto_per_jenis[jenis].append(foto)

    urutan_halaman = [
        ("barang", "FOTO BARANG"), ("kegiatan", "FOTO KEGIATAN"),
        ("sebelum", "FOTO SEBELUM"), ("proses", "FOTO PROSES"), ("sesudah", "FOTO SESUDAH"),
    ]

    # --- KONFIGURASI TAMPILAN ---
    img_height = 70       # Tinggi foto tetap (mm)
    page_margin_x = 10    # Margin kiri/kanan halaman PDF
    max_line_width = pdf.w - (page_margin_x * 2) # Lebar area kerja efektif
    bottom_margin = 20    # Batas bawah halaman sebelum break

    # Fungsi Helper: Render Header Halaman
    def print_header_lampiran(judul_section):
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(0, 8, f"LAMPIRAN {judul_section}", align="C", ln=1)
        pdf.set_font("helvetica", "", 11)
        pdf.cell(0, 6, f"Nomor Kwitansi : {kwitansi.nomor_kwitansi or '-'}", align="C", ln=1)
        pdf.ln(2) # Jarak sedikit setelah header

    for jenis, judul in urutan_halaman:
        daftar_foto = foto_per_jenis.get(jenis, [])
        if not daftar_foto: continue

        # Buat halaman baru untuk setiap jenis kategori (opsional, agar rapi)
        pdf.add_page()
        print_header_lampiran(judul)
        
        # Buffer untuk baris saat ini
        row_buffer = [] 
        current_row_width = 0

        # Fungsi Helper: Flush (Cetak) Buffer ke PDF
        def flush_row(buffer, row_width, current_judul):
            # Cek apakah cukup ruang vertikal?
            if pdf.get_y() + img_height > pdf.h - bottom_margin:
                pdf.add_page()
                print_header_lampiran(current_judul)
            
            # Hitung posisi X awal agar RATA TENGAH
            start_x = (pdf.w - row_width) / 2
            current_x = start_x
            current_y = pdf.get_y()

            for item in buffer:
                # Render Gambar
                if item['valid']:
                    try:
                        pdf.image(str(item['path']), x=current_x, y=current_y, h=img_height)
                    except Exception:
                        # Fallback jika gagal render gambar valid
                        pdf.rect(current_x, current_y, item['width'], img_height)
                else:
                    # Render Placeholder (Kotak Silang/Kosong)
                    pdf.rect(current_x, current_y, item['width'], img_height)
                    pdf.set_xy(current_x, current_y + (img_height/2) - 3)
                    pdf.set_font("helvetica", "I", 7)
                    pdf.multi_cell(item['width'], 3, "Foto Tidak\nDitemukan", align="C")
                
                # Geser X untuk gambar berikutnya (Rapat tanpa spasi)
                current_x += item['width']

            # Pindahkan kursor Y ke bawah untuk baris selanjutnya
            pdf.set_xy(page_margin_x, current_y + img_height)

        # --- LOOP PROSES FOTO ---
        for i, foto in enumerate(daftar_foto):
            foto_path_obj = resolve_foto_path(foto.get("path", ""))
            
            # Default width untuk placeholder jika gambar rusak
            display_width = 50 
            is_valid = False

            # Coba dapatkan dimensi asli gambar untuk hitung lebar proporsional
            if foto_path_obj and foto_path_obj.exists():
                try:
                    with Image.open(foto_path_obj) as img:
                        w_ori, h_ori = img.size
                        aspect_ratio = w_ori / h_ori
                        display_width = img_height * aspect_ratio
                        is_valid = True
                except Exception:
                    is_valid = False
            
            # Siapkan objek data gambar
            img_data = {
                'path': foto_path_obj,
                'width': display_width,
                'valid': is_valid
            }

            # Cek apakah baris penuh?
            if current_row_width + display_width > max_line_width:
                flush_row(row_buffer, current_row_width, judul)
                row_buffer = []     # Kosongkan buffer
                current_row_width = 0 # Reset lebar

            # Masukkan ke buffer
            row_buffer.append(img_data)
            current_row_width += display_width

        # Cetak sisa foto yang ada di buffer terakhir
        if row_buffer:
            flush_row(row_buffer, current_row_width, judul)
            
        # Beri jarak sedikit antar kategori
        pdf.ln(5)

    return pdf.output(dest="S").encode("latin1")

# --- Route Cetak PDF BAST ---
@api_router.get("/kwitansi/bast/{no_bukti}")
async def cetak_bast_by_no_bukti(no_bukti: str, db: AsyncSession = Depends(get_db)):
    # 1. Ambil Data Kwitansi
    result = await db.execute(select(Kwitansi).where(Kwitansi.no_bukti == no_bukti))
    kwitansi = result.scalar_one_or_none()
    if not kwitansi:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # 2. Ambil Pengaturan
    result = await db.execute(select(Pengaturan))
    pengaturan = result.scalar_one_or_none()
    
    # [cite_start]PROTEKSI: Jika tabel pengaturan kosong sama sekali [cite: 1136]
    if not pengaturan:
        class Dummy:
            nama_pemda = nama_sekolah = alamat_sekolah = ""
            nama_kepala_sekolah = nip_kepala_sekolah = ""
            nama_bendahara = nip_bendahara = ""
            nama_pengurus_barang = nip_pengurus_barang = ""
            logo_pemda = logo_sekolah = None
            tempat_surat = tempat_surat = ""
        pengaturan = Dummy()
    
    # 3. Ambil Transaksi
    result = await db.execute(
        select(Transaksi).where(Transaksi.no_bukti == no_bukti).order_by(Transaksi.id)
    )
    transaksi_list = result.scalars().all()

    # 4. Generate PDF menggunakan fungsi Helper (HASIL PASTI SAMA)
    pdf_bytes = generate_bast_pdf(kwitansi, pengaturan, transaksi_list)
    
    filename = f"BAST-{no_bukti}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

@api_router.get("/kwitansi/{id}/pdf")
async def get_kwitansi_pdf(
    id: int,
    db: AsyncSession = Depends(get_db)
):
    # 1. Ambil Data Kwitansi
    stmt = select(Kwitansi).where(Kwitansi.id == id)
    result = await db.execute(stmt)
    kwitansi = result.scalar_one_or_none()
    
    if not kwitansi:
        raise HTTPException(status_code=404, detail="Kwitansi tidak ditemukan")

    # 2. Ambil Pengaturan
    stmt2 = select(Pengaturan).limit(1)
    result2 = await db.execute(stmt2)
    pengaturan = result2.scalar_one_or_none()
    
    if not pengaturan:
        raise HTTPException(status_code=404, detail="Pengaturan belum diisi")

    # 3. Generate Bagian 1: KWITANSI
    try:
        kw_pdf_bytes = generate_kwitansi_pdf_v2(kwitansi, pengaturan)
    except Exception as e:
        logger.error(f"Gagal generate Kwitansi PDF: {e}")
        raise HTTPException(status_code=500, detail="Gagal membuat PDF Kwitansi")

    # 4. Generate Bagian 2: BAST (Jika ada no_bukti)
    bast_pdf_bytes = None
    if kwitansi.no_bukti:
        # Ambil item transaksi untuk tabel BAST
        stmt_trx = select(Transaksi).where(Transaksi.no_bukti == kwitansi.no_bukti).order_by(Transaksi.id)
        res_trx = await db.execute(stmt_trx)
        transaksi_list = res_trx.scalars().all()

        if transaksi_list:
            try:
                # Panggil fungsi generate_bast_pdf yang sudah ada di kode Anda
                bast_pdf_bytes = generate_bast_pdf(kwitansi, pengaturan, transaksi_list)
            except Exception as e:
                # Jika BAST gagal (misal karena logo), catat log tapi jangan hentikan proses (opsional)
                logger.error(f"Gagal generate BAST PDF: {e}")
                # bast_pdf_bytes tetap None

    # 5. Gabungkan PDF (Merge)
    final_buffer = BytesIO()
    
    if bast_pdf_bytes:
        # Jika BAST berhasil dibuat, gabungkan Kwitansi + BAST
        merger = PdfMerger()
        merger.append(BytesIO(kw_pdf_bytes))
        merger.append(BytesIO(bast_pdf_bytes))
        merger.write(final_buffer)
        merger.close()
    else:
        # Jika tidak ada BAST, kirim Kwitansi saja
        final_buffer.write(kw_pdf_bytes)

    final_buffer.seek(0)

    # Nama file output
    filename = f"SPJ_Lengkap_{kwitansi.no_bukti or kwitansi.id}.pdf"

    return StreamingResponse(
        final_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

# Include router in app (harus di akhir setelah semua routes didefinisikan)
app.include_router(api_router)


if __name__ == "__main__":
    import uvicorn
    # Gunakan 127.0.0.1 agar sama persis dengan frontend
    uvicorn.run(app, host="127.0.0.1", port=8000, log_config=None)
